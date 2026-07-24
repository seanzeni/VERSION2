from __future__ import annotations

# Purpose:
#     Shared report helper functions.
#
# Used By:
#     All report generators.
#
# Responsibilities:
#     - Report output folder creation.
#     - Safe release folder naming.
#     - History folder handling.
#     - Common element sorting.
#     - Common CSV helpers.
#
# Notes:
#     This file should not generate report content.

import csv
import os
import shutil
import stat
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from app.core.models import Element


def safe_release_name(
    release: str,
) -> str:
    value = str(release).strip()

    invalid = '\\/:*?"<>| '

    for char in invalid:
        value = value.replace(
            char,
            "_",
        )

    return value


def get_release_folder(
    release: str,
    base_path: str | Path,
) -> Path:
    release_folder = Path(base_path) / safe_release_name(release)

    release_folder.mkdir(
        parents=True,
        exist_ok=True,
    )

    return release_folder


def get_date_folder(
    release: str,
    base_path: str | Path,
) -> Path:
    date_folder = get_date_folder_path(
        release=release,
        base_path=base_path,
    )

    date_folder.mkdir(
        parents=True,
        exist_ok=True,
    )

    return date_folder


def get_date_folder_path(
    release: str,
    base_path: str | Path,
) -> Path:
    return (
        Path(base_path)
        / safe_release_name(release)
        / datetime.now().strftime("%Y-%m-%d")
    )


def make_writable(
    file_path: Path,
) -> None:
    if file_path.exists():
        os.chmod(
            file_path,
            stat.S_IREAD | stat.S_IWRITE,
        )


def make_read_only(
    file_path: Path,
) -> None:
    if file_path.exists():
        os.chmod(
            file_path,
            stat.S_IREAD,
        )


def get_unique_path(
    path: Path,
) -> Path:
    if not path.exists():
        return path

    timestamp = datetime.now().strftime("%H%M%S")
    candidate = path.with_name(f"{path.stem}_{timestamp}{path.suffix}")

    counter = 1
    while candidate.exists():
        candidate = path.with_name(f"{path.stem}_{timestamp}_{counter}{path.suffix}")
        counter += 1

    return candidate


def build_report_file_prefix(
    release: str,
    move_date: str | object | None,
) -> str:
    release_name = safe_release_name(release).upper() or "UNKNOWN_RELEASE"
    date_name = safe_release_name(_format_move_date(move_date)).upper()

    return f"{release_name}_{date_name}"


def archive_matching_reports(
    target_folder: Path,
    file_prefix: str,
) -> None:
    history_folder = target_folder / "History"

    history_folder.mkdir(
        exist_ok=True,
    )

    for file_path in target_folder.iterdir():
        if not file_path.is_file():
            continue

        if not file_path.name.upper().startswith(f"{file_prefix.upper()}_"):
            continue

        destination = get_unique_path(
            history_folder / file_path.name,
        )

        try:
            make_writable(file_path)
            shutil.move(
                str(file_path),
                str(destination),
            )
            make_read_only(destination)
        except PermissionError as exc:
            raise PermissionError(
                f"Unable to archive {file_path.name}. Close the file if it is open and try again."
            ) from exc


def archive_existing_reports(
    target_folder: Path,
) -> None:
    history_folder = target_folder / "History"

    history_folder.mkdir(
        exist_ok=True,
    )

    for file_path in target_folder.iterdir():
        if not file_path.is_file():
            continue

        destination = get_unique_path(
            history_folder / file_path.name,
        )

        try:
            make_writable(file_path)
            shutil.move(
                str(file_path),
                str(destination),
            )
            make_read_only(destination)
        except PermissionError as exc:
            raise PermissionError(
                f"Unable to archive {file_path.name}. Close the file if it is open and try again."
            ) from exc


def prefix_report_files(
    files: list[Path],
    file_prefix: str,
) -> list[Path]:
    renamed: list[Path] = []

    for path in files:
        if not path.exists():
            continue

        if path.name.upper().startswith(f"{file_prefix.upper()}_"):
            renamed.append(path)
            continue

        destination = get_unique_path(path.with_name(f"{file_prefix}_{path.name}"))

        make_writable(path)
        path.replace(destination)
        make_read_only(destination)
        renamed.append(destination)

    return renamed


def _format_move_date(
    move_date: str | object | None,
) -> str:
    if move_date is None:
        return "UNKNOWN_DATE"

    if hasattr(move_date, "strftime"):
        return move_date.strftime("%Y-%m-%d")

    value = str(move_date).strip()
    if not value:
        return "UNKNOWN_DATE"

    return value[:10]


def sort_elements(
    elements: list[Element],
) -> list[Element]:
    """
    Final report ordering rule.

    1. Errors
    2. Warnings
    3. Everything else

    Within each section:
        Element name alphabetical.
    """

    severity_rank = {
        "ERROR": 0,
        "WARNING": 1,
        "INFO": 2,
        "OK": 3,
    }

    return sorted(
        (element for element in elements if element.visible),
        key=lambda element: (
            severity_rank.get(
                getattr(
                    element.severity,
                    "value",
                    "OK",
                ),
                99,
            ),
            element.element.upper(),
        ),
    )


def export_csv(
    output_path: Path,
    headers: list[str],
    rows: list[list[str]],
) -> None:
    output_path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )
    make_writable(output_path)

    with output_path.open(
        "w",
        newline="",
        encoding="utf-8",
    ) as file:
        writer = csv.writer(file)

        writer.writerow(headers)

        writer.writerows(rows)

    make_read_only(output_path)


def export_xlsx(
    output_path: Path,
    sheets: dict[str, tuple[list[str], list[list[object]]]],
) -> None:
    output_path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )
    make_writable(output_path)

    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    for sheet_name, (headers, rows) in sheets.items():
        worksheet = workbook.create_sheet(
            title=_safe_sheet_name(sheet_name),
        )
        worksheet.append(headers)

        for cell in worksheet[1]:
            cell.font = Font(bold=True)

        for row in rows:
            worksheet.append(row)

        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        _autosize_columns(worksheet)

    workbook.save(output_path)
    make_read_only(output_path)


def _safe_sheet_name(
    sheet_name: str,
) -> str:
    safe_name = str(sheet_name).strip() or "Sheet"

    for char in "[]:*?/\\":
        safe_name = safe_name.replace(char, "_")

    return safe_name[:31]


def _autosize_columns(
    worksheet,
) -> None:
    for column_cells in worksheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        column_letter = get_column_letter(column_cells[0].column)
        worksheet.column_dimensions[column_letter].width = min(
            max(max_length + 2, 10),
            60,
        )
