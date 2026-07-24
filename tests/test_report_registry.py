from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.core.app_state import AppState
from app.core.models import Element
from app.core.models import InventoryIssue
from app.core.models import ReleaseEffort
from app.core.models import ScheduleStatus
from app.reports.report_registry import ReportRegistry
from app.reports.report_utils import make_writable
from app.services.mainframe_location_service import MainframeLocationService
from app.services.stats_service import StatsService


def make_registry(
    location_service=None,
) -> ReportRegistry:
    stats = StatsService(
        workload_settings={
            "default_thread_count": 1,
            "type_categories": {
                "APS": ["OAPS"],
                "COBOL": ["OCOB"],
                "JCL": ["JCL"],
                "LINKDECK": ["LINK"],
                "NON_COMPILE": ["PROC"],
                "X_ELEMENTS": ["XML"],
            },
            "types_per_hour_per_thread": {
                "PROD": {
                    "APS": 5,
                    "COBOL": 10,
                    "JCL": 30,
                    "LINKDECK": 10,
                    "NON_COMPILE": 20,
                    "X_ELEMENTS": 10,
                }
            },
        }
    )
    return ReportRegistry(
        stats_service=stats,
        location_service_provider=lambda: location_service,
        system_region_lookup_provider=lambda: {"SYSTEM01": "DV9"},
        effort_testing_region_lookup_provider=lambda effort_ids: {
            effort_id: [
                (
                    "DV9",
                    "2026-06-01",
                )
            ]
            for effort_id in effort_ids
        },
    )


def make_state() -> AppState:
    state = AppState(
        release="REL1",
        mode="PROD",
        thread_count=1,
    )
    state.loaded_elements = [
        Element(
            release="REL1",
            project="ABC",
            element="PGM001",
            type="OCOB",
            source_row={
                "Application": "APP",
                "Area": "AREA",
                "Package": "PKG",
                "Service": "SVC",
                "Submitter": "USER1",
            },
        )
    ]
    state.all_release_elements = list(state.loaded_elements)
    return state


def make_location_line(
    element: str,
    type_: str,
    env: str,
    version: str,
    ccid: str,
) -> str:
    fields = [
        (element, 8),
        (type_, 8),
        ("SYSTEM01", 8),
        ("SUB1", 4),
        (env, 5),
        ("2026/06/22", 10),
        ("12:00:00:00", 11),
        (version, 5),
        ("USER01", 8),
        (ccid, 7),
        ("COMMENTS", 40),
    ]
    return " ".join(value.ljust(width)[:width] for value, width in fields)


def make_location_service(
    tmp_path: Path,
) -> MainframeLocationService:
    path = tmp_path / "locations.txt"
    path.write_text(
        "\n".join(
            [
                make_location_line("PGM001", "OCOB", "PROD1", "01.03", "CCID00"),
                make_location_line("PGM001", "OCOB", "SYST1", "01.02", "CCID01"),
                make_location_line("PGM001", "OCOB", "QUAL1", "01.01", "CCID02"),
            ]
        ),
        encoding="cp1252",
    )
    return MainframeLocationService().load_file(path)


def test_get_names_contains_core_reports() -> None:
    """Verifies the registry exposes every report shown in Report Center."""
    names = make_registry().get_names()

    assert "Effort Summary Report" in names
    assert "HIPPA Listeners" in names
    assert "Issues Report" in names
    assert "OSG/COPS Report" in names
    assert "ODS Elements" in names
    assert "Release Estimate Report" in names
    assert "Release Inventory Report" in names
    assert "Resync Report" in names


def test_unknown_report_raises_key_error() -> None:
    """Verifies unknown report names fail fast."""
    with pytest.raises(KeyError):
        make_registry().create("Missing Report")


def test_generate_issues_report_csv(tmp_path: Path) -> None:
    """Verifies Issues Report can generate CSV output."""
    output = make_registry().generate(
        "Issues Report",
        "csv",
        make_state(),
        tmp_path,
        True,
    )

    assert output is not None and output.exists()
    make_writable(output)


def test_generate_issues_report_xlsx_includes_glossary_sheet(
    tmp_path: Path,
) -> None:
    """Verifies Issues Report XLSX includes both issue and glossary sheets."""
    output = make_registry().generate(
        "Issues Report",
        "xlsx",
        make_state(),
        tmp_path,
        True,
    )

    assert output is not None and output.suffix == ".xlsx" and output.exists()

    workbook = load_workbook(output, read_only=True)
    assert "Issues Report" in workbook.sheetnames
    assert "Issues Report Status Glossary" in workbook.sheetnames
    workbook.close()
    make_writable(output)


def test_generate_effort_summary_xlsx(tmp_path: Path) -> None:
    """Verifies Effort Summary Report XLSX has separate workbook sheets."""
    output = make_registry().generate(
        "Effort Summary Report",
        "xlsx",
        make_state(),
        tmp_path,
        True,
    )

    assert output is not None and output.suffix == ".xlsx" and output.exists()

    workbook = load_workbook(output, read_only=True)
    assert "Summary" in workbook.sheetnames
    assert "Inventory" in workbook.sheetnames
    assert "Information" in workbook.sheetnames
    workbook.close()
    make_writable(output)


def test_generate_pdf_not_implemented() -> None:
    """Verifies reports without PDF support raise a clear error."""
    with pytest.raises(NotImplementedError):
        make_registry().generate(
            "Issues Report",
            "pdf",
            AppState(),
            Path("."),
            True,
        )


def test_generate_effort_summary_pdf(tmp_path: Path) -> None:
    """Verifies Effort Summary Report can generate PDF output."""
    output = make_registry().generate(
        "Effort Summary Report",
        "pdf",
        make_state(),
        tmp_path,
        True,
    )

    assert output is not None and output.suffix == ".pdf" and output.exists()
    make_writable(output)


def test_generate_release_estimate_pdf(tmp_path: Path) -> None:
    """Verifies Release Estimate Report can generate PDF output."""
    state = make_state()
    state.effort_dates = {"ABC": "2026-06-22"}

    output = make_registry().generate(
        "Release Estimate Report",
        "pdf",
        state,
        tmp_path,
        True,
    )

    assert output is not None and output.suffix == ".pdf" and output.exists()
    make_writable(output)


def test_generate_release_inventory_pdf(tmp_path: Path) -> None:
    """Verifies Release Inventory Report can generate PDF output."""
    state = make_state()
    state.inventory_issues = [
        InventoryIssue(
            release="REL1",
            effort_id="ABC",
            issue_type=ScheduleStatus.SQL_EXPECTED_INVENTORY_MISSING,
            reason="Missing inventory",
        )
    ]
    state.release_efforts = [ReleaseEffort(effort_id="ABC")]

    output = make_registry().generate(
        "Release Inventory Report",
        "pdf",
        state,
        tmp_path,
        True,
    )

    assert output is not None and output.suffix == ".pdf" and output.exists()
    make_writable(output)


def test_generate_osg_cops_pdf(tmp_path: Path) -> None:
    """Verifies OSG/COPS Report can generate PDF output."""
    output = make_registry().generate(
        "OSG/COPS Report",
        "pdf",
        make_state(),
        tmp_path,
        True,
    )

    assert output is not None and output.suffix == ".pdf" and output.exists()
    make_writable(output)


def test_generate_osg_cops_csv_is_not_supported(tmp_path: Path) -> None:
    """OSG/COPS intentionally supports only XLSX and PDF."""
    with pytest.raises(NotImplementedError):
        make_registry().generate(
            "OSG/COPS Report",
            "csv",
            make_state(),
            tmp_path,
            True,
        )


def test_generate_resync_report_csv(tmp_path: Path) -> None:
    """Verifies Resync Report can generate CSV output from NDVR data."""
    output = make_registry(make_location_service(tmp_path)).generate(
        "Resync Report",
        "csv",
        make_state(),
        tmp_path,
        True,
    )

    assert output is not None and output.exists()
    assert "Found PGM001 OCOB in system region DV9" in output.read_text(
        encoding="utf-8"
    )
    make_writable(output)


def test_generate_resync_report_xlsx(tmp_path: Path) -> None:
    """Verifies Resync Report can generate XLSX output."""
    output = make_registry(make_location_service(tmp_path)).generate(
        "Resync Report",
        "xlsx",
        make_state(),
        tmp_path,
        True,
    )

    assert output is not None and output.suffix == ".xlsx" and output.exists()
    make_writable(output)


def test_generate_resync_report_pdf(tmp_path: Path) -> None:
    """Verifies Resync Report can generate PDF output."""
    output = make_registry(make_location_service(tmp_path)).generate(
        "Resync Report",
        "pdf",
        make_state(),
        tmp_path,
        True,
    )

    assert output is not None and output.suffix == ".pdf" and output.exists()
    make_writable(output)


def test_resync_report_uses_all_release_elements_for_same_move_date(
    tmp_path: Path,
) -> None:
    """Resync uses all effort rows for the move date, even blocked table rows."""
    state = make_state()
    state.mode = "PROD"
    state.effort_dates = {
        "ABC": "2026-06-22",
        "XYZ": "2026-06-22",
        "LATER": "2026-07-01",
    }
    blocked = Element(
        release="REL1",
        project="XYZ",
        element="PGM002",
        type="OCOB",
        selected=False,
        selectable=False,
        visible=False,
        source_row={
            "Application": "APP",
            "Submitter": "USER2",
        },
    )
    later = Element(
        release="REL1",
        project="LATER",
        element="PGM003",
        type="OCOB",
        source_row={
            "Application": "APP",
            "Submitter": "USER3",
        },
    )
    state.all_release_elements = [
        *state.loaded_elements,
        blocked,
        later,
    ]
    path = tmp_path / "locations.txt"
    path.write_text(
        "\n".join(
            [
                make_location_line("PGM001", "OCOB", "PROD1", "01.03", "ABC000"),
                make_location_line("PGM001", "OCOB", "SYST1", "01.02", "ABC000"),
                make_location_line("PGM002", "OCOB", "PROD1", "01.03", "XYZ000"),
                make_location_line("PGM002", "OCOB", "SYST1", "01.02", "XYZ000"),
                make_location_line("PGM003", "OCOB", "PROD1", "01.03", "LATER0"),
                make_location_line("PGM003", "OCOB", "SYST1", "01.02", "LATER0"),
            ]
        ),
        encoding="cp1252",
    )
    location_service = MainframeLocationService().load_file(path)

    output = make_registry(location_service).generate(
        "Resync Report",
        "csv",
        state,
        tmp_path,
        True,
    )

    report_text = output.read_text(encoding="utf-8")
    assert "PGM001" in report_text
    assert "PGM002" in report_text
    assert "PGM003" not in report_text
    make_writable(output)
