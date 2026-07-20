from __future__ import annotations

import importlib.util
import os
import sys
import subprocess
from datetime import date
from pathlib import Path

import pandas as pd
import requests
from openpyxl import load_workbook


SCRIPT_PATH = Path(__file__).resolve().parents[1] / "scripts" / "fixp_daily_compare.py"
SPEC = importlib.util.spec_from_file_location("fixp_daily_compare", SCRIPT_PATH)
assert SPEC is not None and SPEC.loader is not None
fixp_module = importlib.util.module_from_spec(SPEC)
sys.modules["fixp_daily_compare"] = fixp_module
SPEC.loader.exec_module(fixp_module)


class FakePersonResolver:
    def resolve(
        self,
        criteria: str,
    ):
        return {
            "USER01": fixp_module.PersonDirectoryInfo(
                name="User One",
                employee_id="ADU1",
                supervisor_id="MGR1",
            ),
            "USER02": fixp_module.PersonDirectoryInfo(
                name="User Two",
                employee_id="ADU2",
                supervisor_id="MGR2",
            ),
            "TL01": fixp_module.PersonDirectoryInfo(
                name="Taylor One",
                employee_id="ADT1",
            ),
            "TL02": fixp_module.PersonDirectoryInfo(
                name="Taylor Two",
                employee_id="ADT2",
            ),
            "TL03": fixp_module.PersonDirectoryInfo(
                name="Taylor Three",
                employee_id="ADT3",
            ),
        }.get(
            criteria,
            fixp_module.PersonDirectoryInfo(name=criteria),
        )

    def resolve_name(
        self,
        ad_id: str,
    ) -> str:
        return {
            "MGR1": "Manager One",
            "MGR2": "Manager Two",
        }.get(
            ad_id,
            ad_id,
        )


class FakeNameResolver:
    def resolve_name(
        self,
        ad_id: str,
    ) -> str:
        return {
            "ADU2": "User Two",
            "MGR2": "Manager Two",
        }.get(
            ad_id,
            ad_id,
        )


def make_settings(
    tmp_path: Path,
    inventory_path: Path,
    fixp_folder: Path,
    ndvr_folder: Path | None = None,
) -> dict:
    return {
        "files": {
            "default_fixp_folder": str(fixp_folder),
            "default_input_file": str(inventory_path),
            "default_ndvr_file": str(ndvr_folder or tmp_path / "missing-ndvr"),
            "default_output_folder": str(tmp_path / "output"),
        },
        "required_columns": [
            "Release",
            "DSN ID",
            "Project",
            "Element",
            "Type",
        ],
    }


def make_fixp_line(
    element: str,
    type_: str,
    system: str,
    subsystem: str,
    env: str,
    generated_date: str,
    version: str,
    user: str,
    ccid: str,
    comments: str = "COMMENTS",
) -> str:
    fields = [
        (element, 8),
        (type_, 8),
        (system, 8),
        (subsystem, 4),
        (env, 5),
        (generated_date, 10),
        ("12:00:00:00", 11),
        (version, 5),
        (user, 8),
        (ccid, 7),
        (comments, 40),
        ("00000", 5),
        ("PKG001", 16),
    ]
    return " ".join(value.ljust(width)[:width] for value, width in fields)


def write_inventory(
    tmp_path: Path,
) -> Path:
    """Creates inventory rows used to enrich FIXP comparison output."""
    path = tmp_path / "inventory.xlsx"
    pd.DataFrame(
        [
            {
                "Release": "2026/07 release",
                "DSN ID": "TL01OWNER",
                "Project": "ABC",
                "Element": "SAME001",
                "Type": "OCOB",
            },
            {
                "Release": "2026/08 release",
                "DSN ID": "TL02OWNER",
                "Project": "XYZ",
                "Element": "MOD001",
                "Type": "OCOB",
            },
            {
                "Release": "2026/09 release",
                "DSN ID": "TL03OWNER",
                "Project": "KEEP",
                "Element": "KEEP001",
                "Type": "OCOB",
            },
        ]
    ).to_excel(path, index=False)
    return path


def write_fixp_files(
    tmp_path: Path,
) -> Path:
    """Creates two-day FIXP snapshots where one row changes and one is deleted."""
    folder = tmp_path / "fixp"
    folder.mkdir()

    (folder / "FIXP-20260714_080000.txt").write_text(
        "\n".join(
            [
                make_fixp_line(
                    "SAME001",
                    "OCOB",
                    "SYSTEM01",
                    "SUB1",
                    "FIXP1",
                    "2026/07/14",
                    "01.01",
                    "USER01",
                    "CCID01",
                ),
                make_fixp_line(
                    "MOD001",
                    "OCOB",
                    "SYSTEM01",
                    "SUB1",
                    "FIXP1",
                    "2026/07/14",
                    "01.01",
                    "USER01",
                    "CCID01",
                ),
                make_fixp_line(
                    "KEEP001",
                    "OCOB",
                    "SYSTEM01",
                    "SUB1",
                    "FIXP1",
                    "2026/07/14",
                    "01.01",
                    "USER01",
                    "CCID01",
                ),
                make_fixp_line(
                    "DROP001",
                    "OCOB",
                    "SYSTEM01",
                    "SUB1",
                    "FIXP1",
                    "2026/07/14",
                    "01.01",
                    "USER01",
                    "CCID01",
                ),
            ]
        ),
        encoding="cp1252",
    )

    (folder / "FIXP-20260715_080000.txt").write_text(
        "\n".join(
            [
                make_fixp_line(
                    "SAME001",
                    "OCOB",
                    "SYSTEM01",
                    "SUB1",
                    "FIXP1",
                    "2026/07/14",
                    "01.01",
                    "USER01",
                    "CCID01",
                ),
                make_fixp_line(
                    "MOD001",
                    "OCOB",
                    "SYSTEM01",
                    "SUB1",
                    "FIXP1",
                    "2026/07/15",
                    "01.02",
                    "USER02",
                    "CCID02",
                ),
                make_fixp_line(
                    "KEEP001",
                    "OCOB",
                    "SYSTEM01",
                    "SUB1",
                    "FIXP1",
                    "2026/07/15",
                    "01.01",
                    "USER01",
                    "CCID01",
                ),
            ]
        ),
        encoding="cp1252",
    )

    (folder / "FIXP-20260715_100000.txt").write_text(
        make_fixp_line(
            "SAME001",
            "OCOB",
            "SYSTEM01",
            "SUB1",
            "FIXP1",
            "2026/07/14",
            "01.01",
            "USER01",
            "CCID99",
        ),
        encoding="cp1252",
    )

    return folder


def write_ndvr_files(
    tmp_path: Path,
) -> Path:
    """Creates latest NDVR inventory data used to add FIXP remarks."""
    folder = tmp_path / "ndvr"
    folder.mkdir()
    older_file = folder / "NDVR-older.txt"
    latest_file = folder / "NDVR-latest.txt"

    older_file.write_text(
        make_fixp_line(
            "SAME001",
            "OCOB",
            "SYSTEM01",
            "SUB1",
            "PROD1",
            "2026/07/14",
            "01.01",
            "USER01",
            "CCID01",
        ),
        encoding="cp1252",
    )
    latest_file.write_text(
        "\n".join(
            [
                make_fixp_line(
                    "SAME001",
                    "OCOB",
                    "SYSTEM01",
                    "SUB1",
                    "PROD1",
                    "2026/07/16",
                    "01.01",
                    "USER01",
                    "CCID99",
                ),
                make_fixp_line(
                    "KEEP001",
                    "OCOB",
                    "SYSTEM01",
                    "SUB1",
                    "QUAL1",
                    "2026/07/20",
                    "01.01",
                    "USER01",
                    "CCID99",
                ),
            ]
        ),
        encoding="cp1252",
    )
    os.utime(older_file, (1_000, 1_000))
    os.utime(latest_file, (2_000, 2_000))

    return folder


def test_fixp_daily_compare_builds_expected_rows(
    tmp_path: Path,
) -> None:
    """Verifies day-over-day FIXP rows and inventory references."""
    inventory_path = write_inventory(tmp_path)
    fixp_folder = write_fixp_files(tmp_path)
    ndvr_folder = write_ndvr_files(tmp_path)
    report = fixp_module.FixpDailyCompare(
        settings=make_settings(tmp_path, inventory_path, fixp_folder, ndvr_folder),
        base_dir=tmp_path,
        person_resolver=FakePersonResolver(),
    )

    rows = report.build_rows(date(2026, 7, 15))
    statuses = {row[4]: row[0] for row in rows}

    assert statuses == {
            "DROP001": "deleted",
            "KEEP001": "modified",
            "MOD001": "modified",
            "SAME001": "no change",
        }
    assert next(row[7] for row in rows if row[4] == "SAME001") == "CCID99"
    assert next(row[8] for row in rows if row[4] == "MOD001") == "XYZ"
    assert next(row[9] for row in rows if row[4] == "MOD001") == "User Two"
    assert next(row[10] for row in rows if row[4] == "MOD001") == "Manager Two"
    assert next(row[11] for row in rows if row[4] == "MOD001") == (
        "2026/08 release-XYZ-Taylor Two"
    )
    assert next(row[12] for row in rows if row[4] == "SAME001") == (
        "Newer version in PROD"
    )
    assert next(row[12] for row in rows if row[4] == "KEEP001") == ""
    assert next(row[6] for row in rows if row[4] == "DROP001") == "14-Jul-26"
    assert next(row[9] for row in rows if row[4] == "DROP001") == "User One"


def test_fixp_daily_compare_writes_xlsx(
    tmp_path: Path,
) -> None:
    """Verifies the standalone FIXP report writes one stable latest workbook."""
    inventory_path = write_inventory(tmp_path)
    fixp_folder = write_fixp_files(tmp_path)
    report = fixp_module.FixpDailyCompare(
        settings=make_settings(tmp_path, inventory_path, fixp_folder),
        base_dir=tmp_path,
        person_resolver=FakePersonResolver(),
    )

    output_files = report.run(date(2026, 7, 15))

    assert len(output_files) == 1
    assert output_files[0].name == "fixp1-daily-analysis.xlsx"
    assert output_files[0].parent.name == "FIXP Daily Compare"
    workbook = load_workbook(output_files[0], read_only=True)
    assert workbook.sheetnames == ["FIXP Compare"]
    worksheet = workbook["FIXP Compare"]
    headers = [cell.value for cell in next(worksheet.iter_rows(max_row=1))]
    assert headers[8:13] == [
        "Inventory CCIDs",
        "Owner",
        "Manager",
        "Inventory",
        "Remarks",
    ]
    workbook.close()


def test_fixp_daily_compare_defaults_to_latest_two_file_dates(
    tmp_path: Path,
) -> None:
    """Verifies default comparison uses the latest two FIXP file dates available."""
    inventory_path = write_inventory(tmp_path)
    fixp_folder = write_fixp_files(tmp_path)
    (fixp_folder / "FIXP-20260720_080000.txt").write_text(
        make_fixp_line(
            "KEEP001",
            "OCOB",
            "SYSTEM01",
            "SUB1",
            "FIXP1",
            "2026/07/20",
            "01.02",
            "USER01",
            "CCID20",
        ),
        encoding="cp1252",
    )
    report = fixp_module.FixpDailyCompare(
        settings=make_settings(tmp_path, inventory_path, fixp_folder),
        base_dir=tmp_path,
        person_resolver=FakePersonResolver(),
    )

    output_files = report.run(None)

    assert output_files[0].name == "fixp1-daily-analysis.xlsx"
    rows = report.build_rows(None)
    statuses = {row[4]: row[0] for row in rows}
    assert statuses["KEEP001"] == "modified"
    assert statuses["SAME001"] == "deleted"


def test_fixp_daily_compare_archives_previous_latest_file(
    tmp_path: Path,
) -> None:
    """Verifies an existing latest workbook is renamed before replacement."""
    inventory_path = write_inventory(tmp_path)
    fixp_folder = write_fixp_files(tmp_path)
    output_folder = tmp_path / "output"
    latest_folder = output_folder / "FIXP Daily Compare"
    latest_folder.mkdir(
        parents=True,
    )
    previous_latest = latest_folder / "fixp1-daily-analysis.xlsx"
    previous_latest.write_text(
        "old workbook",
        encoding="utf-8",
    )
    previous_latest.chmod(0o444)
    report = fixp_module.FixpDailyCompare(
        settings=make_settings(tmp_path, inventory_path, fixp_folder),
        base_dir=tmp_path,
        output_folder=output_folder,
        person_resolver=FakePersonResolver(),
    )

    output_files = report.run(date(2026, 7, 15))

    archived_files = sorted(latest_folder.glob("fixp1-daily-analysis - *.xlsx"))
    assert output_files == [previous_latest]
    assert len(archived_files) == 1
    assert archived_files[0].read_text(encoding="utf-8") == "old workbook"
    assert previous_latest.exists()


def test_person_api_resolver_uses_api_ids_and_ad_names(
    monkeypatch,
) -> None:
    """Verifies API IDs are translated to display names through AD lookup."""

    class FakeResponse:
        def raise_for_status(
            self,
        ) -> None:
            return None

        def json(
            self,
        ) -> dict:
            return {"employeeId": "ADU2", "supervisorId": "MGR2"}

    def fake_get(
        url,
        timeout,
        verify,
    ):
        assert url == "https://people.example/api?criteria=USER02"
        assert timeout == 10
        assert verify is False
        return FakeResponse()

    monkeypatch.setattr(
        fixp_module.requests,
        "get",
        fake_get,
    )
    resolver = fixp_module.PersonApiResolver(
        "https://people.example/api",
        name_resolver=FakeNameResolver(),
    )

    person = resolver.resolve("USER02")

    assert person.name == "User Two"
    assert person.employee_id == "ADU2"
    assert person.supervisor_id == "MGR2"
    assert resolver.resolve_name(person.supervisor_id) == "Manager Two"


def test_person_api_resolver_falls_back_to_powershell(
    monkeypatch,
) -> None:
    """Verifies Windows-native API lookup is tried when Python URL open fails."""

    def fake_get(
        url,
        timeout,
        verify,
    ):
        raise requests.exceptions.ConnectionError(
            "corporate proxy requires Windows auth"
        )

    def fake_run(
        command,
        capture_output,
        check,
        encoding,
        errors,
        timeout,
    ):
        assert "Invoke-RestMethod" in command[-1]
        assert "-UseDefaultCredentials" in command[-1]
        return subprocess.CompletedProcess(
            args=command,
            returncode=0,
            stdout='[{"employeeId":"ADU2","supervisorId":"MGR2"}]',
            stderr="",
        )

    monkeypatch.setattr(
        fixp_module.requests,
        "get",
        fake_get,
    )
    monkeypatch.setattr(
        fixp_module.subprocess,
        "run",
        fake_run,
    )
    resolver = fixp_module.PersonApiResolver(
        "https://people.example/api",
        name_resolver=FakeNameResolver(),
    )

    person = resolver.resolve("USER02")

    assert person.name == "User Two"
    assert person.employee_id == "ADU2"
    assert person.supervisor_id == "MGR2"


def test_person_api_resolver_ignores_ssl_errors_with_powershell_fallback(
    monkeypatch,
) -> None:
    """Verifies SSL handshake errors do not stop person lookup."""

    def fake_get(
        url,
        timeout,
        verify,
    ):
        raise requests.exceptions.SSLError("sslv3 alert handshake failure")

    def fake_run(
        command,
        capture_output,
        check,
        encoding,
        errors,
        timeout,
    ):
        return subprocess.CompletedProcess(
            args=command,
            returncode=0,
            stdout='{"employeeId":"ADU2","supervisorId":"MGR2"}',
            stderr="",
        )

    monkeypatch.setattr(
        fixp_module.requests,
        "get",
        fake_get,
    )
    monkeypatch.setattr(
        fixp_module.subprocess,
        "run",
        fake_run,
    )
    resolver = fixp_module.PersonApiResolver(
        "https://people.example/api",
        name_resolver=FakeNameResolver(),
    )

    person = resolver.resolve("USER02")

    assert person.name == "User Two"
    assert person.employee_id == "ADU2"
    assert person.supervisor_id == "MGR2"


def test_person_api_resolver_supports_configured_criteria_url() -> None:
    """Verifies settings can include the criteria placeholder directly."""
    resolver = fixp_module.PersonApiResolver(
        "https://people.example/search?criteria=xxxxx",
        name_resolver=FakeNameResolver(),
    )

    assert resolver._lookup_request_url("USER02") == (
        "https://people.example/search?criteria=USER02"
    )


def test_person_api_resolver_supports_trailing_criteria_url() -> None:
    """Verifies settings can end with an empty criteria query parameter."""
    resolver = fixp_module.PersonApiResolver(
        "https://people.example/search?criteria=",
        name_resolver=FakeNameResolver(),
    )

    assert resolver._lookup_request_url("USER02") == (
        "https://people.example/search?criteria=USER02"
    )


def test_parse_target_date_defaults_to_previous_day() -> None:
    """Verifies an explicit CLI date is parsed as the requested report date."""
    assert fixp_module.parse_target_date("2026-07-19", today=date(2026, 7, 20)) == date(
        2026,
        7,
        19,
    )


def test_parse_target_date_returns_none_for_default_file_window() -> None:
    """Verifies no CLI date allows the report to use latest available files."""
    assert fixp_module.parse_target_date(None, today=date(2026, 7, 20)) is None
