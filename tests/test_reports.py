from __future__ import annotations
from pathlib import Path
import csv
from openpyxl import load_workbook
from app.core.models import ArchiveStatus
from app.core.models import Element
from app.core.models import InventoryIssue
from app.core.models import InventoryStatus
from app.core.models import LocationStatus
from app.core.models import MovementStatus
from app.core.models import ReleaseEffort
from app.core.models import ScheduleStatus
from app.reports.effort_summary_report import EffortSummaryReport
from app.reports.issues_report import IssuesReport
from app.reports.osg_cops_report import OsgCopsReport
from app.reports.report_utils import make_writable
from app.reports.release_estimate_report import ReleaseEstimateReport
from app.reports.release_inventory_report import ReleaseInventoryReport
from app.services.stats_service import StatsService

def make_stats_service() -> StatsService:
    return StatsService(workload_settings={"default_thread_count":1,"type_categories":{"JCL":["JCL"],"NON_COMPILE":["PROC"],"COBOL":["OCOB"],"APS":["OAPS"],"X_ELEMENTS":["XML"],"LINKDECK":["LINK"]},"types_per_hour_per_thread":{"PROD":{"JCL":30,"NON_COMPILE":20,"COBOL":10,"APS":5,"X_ELEMENTS":10,"LINKDECK":10}}})

def make_element(name='OPGM001', type_='OCOB', project='ABC', selected=True, visible=True, archive_status=ArchiveStatus.OK, schedule_status=ScheduleStatus.OK) -> Element:
    e=Element(release='REL1', project=project, element=name, type=type_, selected=selected, visible=visible, archive_status=archive_status, schedule_status=schedule_status, source_row={"Submitter":"USER1","Application":"APP","Package":"PKG","Area":"AREA","Service":"SVC"})
    if archive_status != ArchiveStatus.OK or schedule_status != ScheduleStatus.OK:
        e.reasons.append('Test reason')
    return e

def read_csv(path: Path) -> list[dict[str,str]]:
    with path.open('r', encoding='utf-8', newline='') as f:
        return list(csv.DictReader(f))

def test_issues_report_excludes_hidden(tmp_path: Path) -> None:
    """Verifies issues report excludes hidden."""
    output=IssuesReport().generate([make_element(archive_status=ArchiveStatus.POTENTIAL_MISSING_ARCHIVE), make_element(name='OPGM002', visible=False, archive_status=ArchiveStatus.POTENTIAL_MISSING_ARCHIVE)], tmp_path, False)
    rows=read_csv(output)
    assert len(rows)==1 and rows[0]['Element']=='OPGM001'
    glossary=tmp_path / IssuesReport.GLOSSARY_FILE_NAME
    assert glossary.exists()
    assert any(row['Value']=='POTENTIAL_MISSING_ARCHIVE' for row in read_csv(glossary))
    make_writable(output)
    make_writable(glossary)

def test_issues_report_excludes_info_severity(tmp_path: Path) -> None:
    """Verifies issues report only includes warnings and errors."""
    info = make_element()
    info.reasons.append('Informational reason')
    warning = make_element(
        name='OPGM002',
        archive_status=ArchiveStatus.POTENTIAL_MISSING_PROGRAM_MOVE,
    )

    output=IssuesReport().generate([info, warning], tmp_path, False)
    rows=read_csv(output)

    assert [row['Element'] for row in rows] == ['OPGM002']
    assert rows[0]['Severity'] == 'WARNING'
    make_writable(output)

def test_effort_summary_report_generates_rows(tmp_path: Path) -> None:
    """Verifies effort summary report generates rows."""
    output=EffortSummaryReport(make_stats_service()).generate([make_element(project='ABC', type_='OCOB'), make_element(project='ABC', type_='OAPS')], tmp_path, 'PROD', 1)
    rows=read_csv(output)
    assert len(rows)==2
    assert {row['Type'] for row in rows} == {'OCOB', 'OAPS'}
    assert all(row['Project']=='ABC' for row in rows)
    assert all(row['Schedule Status']=='OK' for row in rows)
    make_writable(output)

def test_effort_summary_report_includes_hidden_inventory_details(tmp_path: Path) -> None:
    """Verifies effort summary report includes hidden inventory details."""
    hidden = make_element(
        name='OPGM003',
        project='ABC',
        selected=False,
        visible=False,
    )
    hidden.movement_status = MovementStatus.DO_NOT_MOVE
    hidden.reasons.append('Hidden from this move by marker')

    output=EffortSummaryReport(make_stats_service()).generate([make_element(project='ABC', type_='OCOB'), hidden], tmp_path, 'PROD', 1)
    rows=read_csv(output)

    assert {row['Element'] for row in rows} == {'OPGM001', 'OPGM003'}
    hidden_row=next(row for row in rows if row['Element']=='OPGM003')
    assert hidden_row['Selected']=='False'
    assert hidden_row['Visible']=='False'
    assert hidden_row['Movement Status']=='DO_NOT_MOVE'
    assert hidden_row['Reasons']=='Hidden from this move by marker'
    make_writable(output)

def test_effort_summary_report_includes_hidden_only_effort(tmp_path: Path) -> None:
    """Verifies effort summary report includes hidden only effort."""
    hidden = make_element(
        name='OPGM004',
        project='HIDDEN',
        selected=False,
        visible=False,
    )

    output=EffortSummaryReport(make_stats_service()).generate([hidden], tmp_path, 'PROD', 1)
    rows=read_csv(output)

    assert [row['Project'] for row in rows] == ['HIDDEN']
    assert [row['Element'] for row in rows] == ['OPGM004']
    make_writable(output)

def test_effort_summary_status_count_rows() -> None:
    """Verifies effort summary counts each warning/error status type."""
    overlap = make_element()
    overlap.inventory_status = InventoryStatus.OVERLAP
    missing_archive = make_element(name='OPGM002')
    missing_archive.archive_status = ArchiveStatus.POTENTIAL_MISSING_ARCHIVE
    missing_program = make_element(name='OPGM003')
    missing_program.archive_status = ArchiveStatus.POTENTIAL_MISSING_PROGRAM_MOVE
    not_found = make_element(name='OPGM004')
    not_found.location_status = LocationStatus.NOT_FOUND

    rows = EffortSummaryReport(make_stats_service())._build_status_count_rows(
        [overlap, missing_archive, missing_program, not_found]
    )

    counts = {row[0]: row[1] for row in rows}
    assert counts['Overlaps'] == 1
    assert counts['Missing Archives'] == 1
    assert counts['Missing Programs'] == 1
    assert counts['Not Found'] == 1

def test_release_estimate_report_generates_total(tmp_path: Path) -> None:
    """Verifies release estimate report generates total."""
    output=ReleaseEstimateReport(make_stats_service()).generate([make_element(project='ABC', type_='OCOB')], {'ABC':'2026-06-22'}, tmp_path, 'PROD', 3)
    rows=read_csv(output)
    assert rows[0]['Effort'] == 'ABC'
    assert rows[0]['Move Date'] == '2026-06-22'
    assert rows[-1]['Effort'] == 'TOTAL'
    assert rows[-1]['Thread Count'] == '3'
    make_writable(output)

def test_release_estimate_pdf_rollup_matches_effort_total_rounding() -> None:
    """Verifies PDF estimate sums effort-level estimates like CSV/XLSX total."""
    elements = [
        make_element(project='ABC', type_='OCOB'),
        make_element(project='XYZ', type_='OCOB'),
    ]

    estimate = ReleaseEstimateReport(make_stats_service())._build_effort_rollup(
        elements=elements,
        mode='PROD',
        thread_count=5,
    )

    assert estimate['effort_count'] == 2
    assert estimate['selected_elements'] == 2
    assert estimate['total_minutes'] == 4
    assert estimate['estimated_time'] == '00:04'

def test_release_estimate_forecast_counts_all_movable_rows() -> None:
    """Verifies forecast estimate counts all rows except non-moving markers."""
    selected = make_element(project='ABC', type_='OCOB')
    unselected_issue = make_element(
        name='OPGM002',
        project='ABC',
        type_='OCOB',
        selected=False,
        visible=False,
    )
    already_prod = make_element(name='OPGM003', project='ABC', type_='OCOB')
    already_prod.movement_status = MovementStatus.MARKED_IN_PROD
    do_not_move = make_element(name='OPGM004', project='ABC', type_='OCOB')
    do_not_move.movement_status = MovementStatus.DO_NOT_MOVE
    already_qual = make_element(name='OPGM005', project='ABC', type_='OCOB')
    already_qual.movement_status = MovementStatus.MARKED_IN_QUAL

    estimate = ReleaseEstimateReport(make_stats_service())._build_effort_rollup(
        elements=[
            selected,
            unselected_issue,
            already_prod,
            do_not_move,
            already_qual,
        ],
        mode='PROD',
        thread_count=5,
        count_all_movable_elements=True,
    )

    assert estimate['selected_elements'] == 2
    assert estimate['category_counts']['COBOL'] == 2

def test_release_inventory_report_missing_inventory(tmp_path: Path) -> None:
    """Verifies release inventory report missing inventory."""
    output=ReleaseInventoryReport().generate('REL1','PROD',[],[InventoryIssue(release='REL1', effort_id='ABC', issue_type=ScheduleStatus.SQL_EXPECTED_INVENTORY_MISSING, reason='Missing inventory')],[ReleaseEffort(effort_id='ABC')],tmp_path)
    rows=read_csv(output)
    assert rows[0]['Inventory Status'] == 'Missing Inventory'
    assert 'Thread Count' not in rows[0]
    make_writable(output)

def test_release_inventory_report_withdrawn_inventory_notification(tmp_path: Path) -> None:
    """Verifies release inventory report withdrawn inventory notification."""
    output=ReleaseInventoryReport().generate('REL1','PROD',[make_element(project='ABC')],[],[ReleaseEffort(effort_id='ABC', exit_date='2026-06-24')],tmp_path)
    rows=read_csv(output)
    assert rows[0]['Inventory Status'] == 'Unexpected Inventory'
    assert 'withdrawn' in rows[0]['Reason']
    make_writable(output)

def test_release_inventory_report_excludes_element_only_issues(tmp_path: Path) -> None:
    """Verifies release inventory report excludes non-inventory element issues."""
    element = make_element(project='ABC')
    element.location_status = LocationStatus.NOT_FOUND
    element.reasons = ['Element/type was not found in the expected NDVR location.']

    output=ReleaseInventoryReport().generate('REL1','PROD',[element],[],[ReleaseEffort(effort_id='ABC')],tmp_path)
    rows=read_csv(output)

    assert rows == []
    make_writable(output)

def test_release_inventory_report_uses_sql_release_for_mismatch(tmp_path: Path) -> None:
    """Verifies release inventory mismatch columns use RSET release metadata."""
    element = make_element(
        project='ABC',
        schedule_status=ScheduleStatus.EFFORT_RELEASE_MISMATCH,
    )
    element.reasons = [
        'Element/type was not found in the expected NDVR location.',
        'Inventory identified bundle does not match the RSET bundle for this project. Inventory says project ABC belongs to the release bundle [REL1], but RSET states it belongs to this release [REL2].',
    ]
    element.source_row['_sql_release'] = 'REL2'

    output=ReleaseInventoryReport().generate('REL1','PROD',[element],[],[ReleaseEffort(effort_id='ABC')],tmp_path)
    rows=read_csv(output)

    assert rows[0]['Inventory Status'] == 'Potential Wrong Release'
    assert rows[0]['Expected Release'] == 'REL2'
    assert rows[0]['Inventory Release'] == 'REL1'
    assert 'RSET states it belongs to this release [REL2]' in rows[0]['Reason']
    assert 'expected NDVR location' not in rows[0]['Reason']
    make_writable(output)

def test_osg_cops_report_filters_selected_visible_o_or_x(tmp_path: Path) -> None:
    """Verifies OSG COPS report filters selected visible o or x."""
    output=OsgCopsReport().generate_xlsx([make_element(name='OPGM001', type_='OCOB'), make_element(name='APGM001', type_='JCL'), make_element(name='XPGM001', type_='XCOB', visible=False)], tmp_path, 'PROD')
    workbook = load_workbook(output, read_only=True)
    worksheet = workbook["OSG COPS"]
    rows = list(worksheet.iter_rows(values_only=True))
    assert len(rows) == 2 and rows[1][2] == 'OPGM001'
    workbook.close()
    make_writable(output)


def test_osg_cops_suppresses_archive_when_replacement_is_moving() -> None:
    """A paired replacement prevents the archive row from appearing twice."""
    archive = make_element(name="OPGM001", type_="OAPS")
    archive.source_row["Package"] = "ARCHIVE"
    replacement = make_element(name="OPGM001", type_="OCOB")

    rows = OsgCopsReport([["OAPS", "OCOB"]])._build_rows(
        [archive, replacement],
        "PROD",
    )

    assert [row[3] for row in rows] == ["OCOB"]


def test_osg_cops_keeps_unpaired_package_archive() -> None:
    """An archive without its moving replacement remains visible."""
    archive = make_element(name="OPGM001", type_="OAPS")
    archive.source_row["Package"] = "ARCHIVE"

    rows = OsgCopsReport([["OAPS", "OCOB"]])._build_rows(
        [archive],
        "PROD",
    )

    assert [row[3] for row in rows] == ["OAPS"]
    assert rows[0][5] == "Package archive"


def test_osg_cops_is_empty_outside_prod() -> None:
    """OSG/COPS is only applicable to PROD movement."""
    rows = OsgCopsReport()._build_rows(
        [make_element(name="OPGM001")],
        "QUAL",
    )

    assert rows == []
