from __future__ import annotations

import importlib.util
import os
import sys
from pathlib import Path

from openpyxl import load_workbook

from app.services.mainframe_location_service import MainframeLocationService


SCRIPT_PATH = Path(__file__).resolve().parents[1] / "scripts" / "global_resync_report.py"
SPEC = importlib.util.spec_from_file_location("global_resync_report", SCRIPT_PATH)
assert SPEC is not None and SPEC.loader is not None
resync_module = importlib.util.module_from_spec(SPEC)
sys.modules["global_resync_report"] = resync_module
SPEC.loader.exec_module(resync_module)


def make_settings(
    tmp_path: Path,
    ndvr_folder: Path,
) -> dict:
    return {
        "files": {
            "default_ndvr_file": str(ndvr_folder),
            "default_output_folder": str(tmp_path / "output"),
        }
    }


def make_line(
    element: str,
    type_: str,
    system: str,
    subsystem: str,
    env: str,
    version: str,
    ccid: str,
) -> str:
    fields = [
        (element, 8),
        (type_, 8),
        (system, 8),
        (subsystem, 4),
        (env, 5),
        ("2026/07/16", 10),
        ("12:00:00:00", 11),
        (version, 5),
        ("USER01", 8),
        (ccid, 7),
        ("COMMENTS", 40),
        ("00000", 5),
        ("PKG001", 16),
    ]
    return " ".join(value.ljust(width)[:width] for value, width in fields)


def write_ndvr(
    tmp_path: Path,
) -> Path:
    folder = tmp_path / "ndvr"
    folder.mkdir()
    older = folder / "NDVR-old.txt"
    latest = folder / "NDVR-latest.txt"
    older.write_text(
        make_line("OLD001", "OCOB", "SYSTEM01", "SUB1", "QUAL1", "01.01", "OLD"),
        encoding="cp1252",
    )
    latest.write_text(
        "\n".join(
            [
                make_line("PGM001", "OCOB", "PRIVATE1", "SUB1", "PROD1", "03.00", "PROD"),
                make_line("PGM001", "OCOB", "PRIVATE1", "SUB1", "QUAL1", "02.00", "QUAL"),
                make_line("PGM001", "OCOB", "SYSTEM01", "SUB1", "SYST1", "01.00", "SYST"),
                make_line("PGM001", "OCOB", "SYSTEM02", "SUB1", "UNIT1", "01.50", "UNIT"),
                make_line("PGM001", "OCOB", "SYSTEM03", "SUB1", "FIXP1", "99.99", "FIXP"),
                make_line("EQ001", "OCOB", "SYSTEM01", "SUB1", "QUAL1", "01.01", "Q1"),
                make_line("EQ001", "OCOB", "SYSTEM02", "SUB1", "QUAL1", "01.02", "Q2"),
                make_line("OK001", "OCOB", "SYSTEM01", "SUB1", "SYST1", "02.00", "OK"),
                make_line("OK001", "OCOB", "SYSTEM02", "SUB1", "UNIT1", "03.00", "LOW"),
            ]
        ),
        encoding="cp1252",
    )
    os.utime(older, (1_000, 1_000))
    os.utime(latest, (2_000, 2_000))
    return folder


def test_global_resync_report_builds_rows_for_higher_or_equal_env_versions(
    tmp_path: Path,
) -> None:
    """Verifies global resync scans all lifecycle envs without release context."""
    ndvr_folder = write_ndvr(tmp_path)
    service = MainframeLocationService().load_file(ndvr_folder / "NDVR-latest.txt")
    report = resync_module.GlobalResyncReport(
        settings=make_settings(tmp_path, ndvr_folder),
        base_dir=tmp_path,
    )

    rows = report.build_rows(service)
    row_keys = {
        (
            row[0],
            row[2],
            row[6],
            row[8],
            row[12],
        )
        for row in rows
    }

    assert ("PGM001", "QUAL1", "02.00", "PROD1", "03.00") in row_keys
    assert ("PGM001", "SYST1", "01.00", "QUAL1", "02.00") in row_keys
    assert ("PGM001", "UNIT1", "01.50", "QUAL1", "02.00") in row_keys
    assert ("EQ001", "QUAL1", "01.01", "QUAL1", "01.02") in row_keys
    assert not any(row[8] == "FIXP1" or row[2] == "FIXP1" for row in rows)
    assert not any(row[0] == "OK001" for row in rows)


def test_global_resync_report_writes_xlsx(
    tmp_path: Path,
) -> None:
    """Verifies standalone global resync writes an XLSX workbook."""
    ndvr_folder = write_ndvr(tmp_path)
    report = resync_module.GlobalResyncReport(
        settings=make_settings(tmp_path, ndvr_folder),
        base_dir=tmp_path,
    )

    output_files = report.run()

    assert len(output_files) == 1
    assert output_files[0].name == "Global_Resync_Report.xlsx"
    workbook = load_workbook(output_files[0], read_only=True)
    assert workbook.sheetnames == ["Global Resync"]
    rows = list(workbook["Global Resync"].iter_rows(values_only=True))
    workbook.close()
    assert rows[0][:4] == (
        "Element",
        "Type",
        "Target Env",
        "Target Level",
    )
    assert any(row[0] == "PGM001" for row in rows[1:])
