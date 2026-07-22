from __future__ import annotations

import importlib.util
import os
import sys
from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


SCRIPT_PATH = (
    Path(__file__).resolve().parents[1] / "scripts" / "region_inventory_audit.py"
)
SPEC = importlib.util.spec_from_file_location("region_inventory_audit", SCRIPT_PATH)
assert SPEC is not None and SPEC.loader is not None
audit_module = importlib.util.module_from_spec(SPEC)
sys.modules["region_inventory_audit"] = audit_module
SPEC.loader.exec_module(audit_module)


class FakeAssignmentClient:
    def __init__(self) -> None:
        self.requested_window = None

    def load_region_assignments(
        self,
        start_date: date,
        end_date: date,
        active_date: date,
    ):
        self.requested_window = (
            start_date,
            end_date,
            active_date,
        )
        return [
            audit_module.RegionAssignment(
                bundle_id="2026/07 Release",
                bundle_sequence="100",
                region="DV01",
                system="SYSTEM01",
                effort_id="ABC12345",
                bundle_test_environment="12",
                bundle_prod_imp_date="2026-07-20",
                region_id="DV0 Example",
                region_prefix="DV0",
            ),
            audit_module.RegionAssignment(
                bundle_id="2026/07 Release",
                bundle_sequence="100",
                region="DV01",
                system="SYSTEM01",
                effort_id="XYZ99999",
                bundle_test_environment="12",
                bundle_prod_imp_date="2026-07-20",
                region_id="DV0 Example",
                region_prefix="DV0",
            ),
            audit_module.RegionAssignment(
                bundle_id="2026/08 NPR",
                bundle_sequence="101",
                region="DV02",
                system="SYSTEM02",
                effort_id="NPR12345",
                bundle_test_environment="13",
                bundle_prod_imp_date="2026-08-20",
                region_id="DV0 Example",
                region_prefix="DV0",
            ),
            audit_module.RegionAssignment(
                bundle_id="2026/06 NPR",
                bundle_sequence="99",
                region="DV01",
                system="SYSTEM01",
                effort_id="OLD12345",
                bundle_test_environment="12",
                bundle_prod_imp_date="2026-06-20",
                region_id="DV0 Example",
                region_prefix="DV0",
            ),
        ]


class FakeCursor:
    def __init__(self) -> None:
        self.query = ""
        self.params = ()

    def execute(
        self,
        query,
        *params,
    ) -> None:
        self.query = query
        self.params = params

    def fetchall(self) -> list:
        return []


class FakeConnection:
    def __init__(
        self,
        cursor: FakeCursor,
    ) -> None:
        self._cursor = cursor

    def __enter__(self):
        return self

    def __exit__(
        self,
        exc_type,
        exc,
        traceback,
    ) -> None:
        return None

    def cursor(self) -> FakeCursor:
        return self._cursor


class FakeDbService:
    def __init__(self) -> None:
        self.cursor = FakeCursor()

    def get_connection(self) -> FakeConnection:
        return FakeConnection(self.cursor)


def make_settings(
    tmp_path: Path,
    inventory_path: Path,
    ndvr_folder: Path,
) -> dict:
    return {
        "database": {},
        "files": {
            "default_input_file": str(inventory_path),
            "default_ndvr_file": str(ndvr_folder),
            "default_output_folder": str(tmp_path / "output"),
        },
        "required_columns": [
            "Release",
            "Project",
            "Element",
            "Type",
        ],
    }


def make_ndvr_line(
    element: str,
    type_: str,
    system: str,
    ccid: str,
) -> str:
    fields = [
        (element, 8),
        (type_, 8),
        (system, 8),
        ("SYS1", 4),
        ("DEVL1", 5),
        ("2026/07/14", 10),
        ("12:00:00:00", 11),
        ("01.01", 5),
        ("USER01", 8),
        (ccid, 7),
        ("COMMENTS", 40),
        ("00000", 5),
        ("PKG001", 16),
    ]
    return " ".join(value.ljust(width)[:width] for value, width in fields)


def write_inventory(
    tmp_path: Path,
) -> Path:
    path = tmp_path / "inventory.xlsx"
    pd.DataFrame(
        [
            {
                "Release": "2026/07 release",
                "Project": "ABC12345",
                "Element": "PGM001",
                "Type": "OCOB",
            },
            {
                "Release": "2026/07 release",
                "Project": "OTHER01",
                "Element": "BAD001",
                "Type": "OCOB",
            },
            {
                "Release": "2026/06 NPR",
                "Project": "OLD12345",
                "Element": "OLDPGM",
                "Type": "OCOB",
            },
            {
                "Release": "2026/08 NPR",
                "Project": "NPR12345",
                "Element": "CROSS01",
                "Type": "OCOB",
            },
        ]
    ).to_excel(path, index=False)
    return path


def write_ndvr(
    tmp_path: Path,
) -> Path:
    folder = tmp_path / "ndvr"
    folder.mkdir()
    older = folder / "NDVR-old.txt"
    latest = folder / "NDVR-latest.txt"
    older.write_text(
        make_ndvr_line("OLD001", "OCOB", "SYSTEM01", "ABC"),
        encoding="cp1252",
    )
    latest.write_text(
        "\n".join(
            [
                make_ndvr_line("PGM001", "OCOB", "SYSTEM01", "ABC"),
                make_ndvr_line("MISS001", "OCOB", "SYSTEM01", "XYZ"),
                make_ndvr_line("BAD001", "OCOB", "SYSTEM01", "BAD"),
                make_ndvr_line("OLDPGM", "OCOB", "SYSTEM01", "DIFF"),
                make_ndvr_line("CROSS01", "OCOB", "SYSTEM01", "DIFF"),
                make_ndvr_line("DROP001", "OCOB", "OTHER001", "ABC"),
            ]
        ),
        encoding="cp1252",
    )
    os.utime(older, (1_000, 1_000))
    os.utime(latest, (2_000, 2_000))
    return folder


def test_report_window_uses_current_month_after_inventory_cutoff() -> None:
    """Verifies region audit drops the prior bundle month on the 15th."""
    assert audit_module.report_window(date(2026, 7, 21)) == (
        date(2026, 7, 1),
        date(2026, 11, 1),
    )


def test_report_window_keeps_prior_month_before_inventory_cutoff() -> None:
    """Verifies prior bundle month is included before the 15th."""
    assert audit_module.report_window(date(2026, 7, 14)) == (
        date(2026, 6, 1),
        date(2026, 11, 1),
    )


def test_effort_matches_truncated_ndvr_ccid() -> None:
    """Verifies full effort IDs can match shortened NDVR CCIDs."""
    assert audit_module.effort_matches_ccid("ABC12345", "ABC")
    assert not audit_module.effort_matches_ccid("ABC12345", "XYZ")


def test_sql_region_assignment_client_uses_regions_bridge() -> None:
    """Verifies bundle test environment resolves systems through Regions."""
    db_service = FakeDbService()
    client = audit_module.SqlRegionAssignmentClient(db_service)

    assignments = client.load_region_assignments(
        start_date=date(2026, 6, 1),
        end_date=date(2026, 11, 1),
        active_date=date(2026, 7, 21),
    )

    assert assignments == []
    assert "FROM Regions" in db_service.cursor.query
    assert "GROUP BY" in db_service.cursor.query
    assert "b.Id NOT LIKE '%Special%'" in db_service.cursor.query
    assert "b.Id LIKE '%Release%'" not in db_service.cursor.query
    assert "LEFT(LTRIM(RTRIM(b.Id)), 7) >= ?" in db_service.cursor.query
    assert "b.BundleProdImpDate >= ?" in db_service.cursor.query
    assert "CAST(TestEnvironment AS VARCHAR(50))" in db_service.cursor.query
    assert "LEFT(LTRIM(RTRIM(Id)), 3)" in db_service.cursor.query
    assert "LEFT(LTRIM(RTRIM(mes.Region)), 3)" in db_service.cursor.query
    assert db_service.cursor.params == (
        "2026/06",
        "2026/11",
        date(2026, 7, 21),
        date(2026, 11, 1),
    )


def test_region_inventory_audit_classifies_region_records(
    tmp_path: Path,
) -> None:
    """Verifies approved, warning, and error rows are counted."""
    inventory_path = write_inventory(tmp_path)
    ndvr_folder = write_ndvr(tmp_path)
    assignment_client = FakeAssignmentClient()
    audit = audit_module.RegionInventoryAudit(
        settings=make_settings(tmp_path, inventory_path, ndvr_folder),
        base_dir=tmp_path,
        assignment_client=assignment_client,
    )

    rows = audit.build_rows(today=date(2026, 7, 21))
    statuses = {row.element: row.status for row in rows}

    assert assignment_client.requested_window == (
        date(2026, 7, 1),
        date(2026, 11, 1),
        date(2026, 7, 21),
    )
    assert statuses == {
        "BAD001": audit_module.STATUS_IMPROPER_ACTIVITY,
        "CROSS01": audit_module.STATUS_IMPROPER_ACTIVITY,
        "MISS001": audit_module.STATUS_POTENTIAL_MISSING_INVENTORY,
        "OLDPGM": audit_module.STATUS_APPROVED,
        "PGM001": audit_module.STATUS_APPROVED,
    }
    approved_row = next(row for row in rows if row.element == "PGM001")
    assert approved_row.inventory_effort_ids == ("ABC12345",)
    assert approved_row.approved_effort_ids == ("ABC12345",)
    warning_row = next(row for row in rows if row.element == "MISS001")
    assert warning_row.reason == "Potential missing inventory but effort approved there."
    old_row = next(row for row in rows if row.element == "OLDPGM")
    assert old_row.bundle_id == "2026/06 NPR"
    assert old_row.inventory_effort_ids == ("OLD12345",)
    assert old_row.approved_effort_ids == ("OLD12345",)
    assert not any(
        row.element == "OLDPGM"
        and row.bundle_id == "2026/07 Release"
        and row.status == audit_module.STATUS_IMPROPER_ACTIVITY
        for row in rows
    )
    cross_row = next(row for row in rows if row.element == "CROSS01")
    assert cross_row.region == "DV01"
    assert cross_row.system == "SYSTEM01"
    assert cross_row.bundle_id == "2026/07 Release"
    assert cross_row.inventory_effort_ids == ("NPR12345",)
    assert cross_row.inventory_assigned_bundles == ("2026/08 NPR",)
    assert cross_row.inventory_assigned_region_systems == ("DV02/SYSTEM02",)
    assert (
        "Found CROSS01 OCOB in DV01 / DEVL1 / SYSTEM01 / SYS1"
        in cross_row.reason
    )
    assert "2026/08 NPR in DV02/SYSTEM02" in cross_row.reason


def test_region_inventory_audit_writes_xlsx(
    tmp_path: Path,
) -> None:
    """Verifies the standalone report writes a summary and detail workbook."""
    inventory_path = write_inventory(tmp_path)
    ndvr_folder = write_ndvr(tmp_path)
    audit = audit_module.RegionInventoryAudit(
        settings=make_settings(tmp_path, inventory_path, ndvr_folder),
        base_dir=tmp_path,
        assignment_client=FakeAssignmentClient(),
    )

    output_files = audit.run(today=date(2026, 7, 21))

    assert len(output_files) == 1
    workbook = load_workbook(output_files[0], read_only=True)
    assert workbook.sheetnames == ["Summary", "SQL Assignments", "Detail"]
    summary_rows = list(workbook["Summary"].iter_rows(values_only=True))
    assignment_rows = list(workbook["SQL Assignments"].iter_rows(values_only=True))
    workbook.close()
    assert summary_rows[1:] == [
        ("DV01/SYSTEM01", "2026/06 NPR", 1, 0, 0),
        ("DV01/SYSTEM01", "2026/07 Release", 1, 1, 2),
        ("DV02/SYSTEM02", "2026/08 NPR", 0, 0, 0),
    ]
    assert (
        "2026/07 Release",
        "100",
        "12",
        "2026-07-20",
        "DV0 Example",
        "DV0",
        "DV01",
        "SYSTEM01",
        "ABC12345",
    ) in assignment_rows
