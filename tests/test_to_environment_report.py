from __future__ import annotations

import importlib.util
import os
import sys
from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


SCRIPT_PATH = (
    Path(__file__).resolve().parents[1] / "scripts" / "to_environment_report.py"
)
SPEC = importlib.util.spec_from_file_location("to_environment_report", SCRIPT_PATH)
assert SPEC is not None and SPEC.loader is not None
report_module = importlib.util.module_from_spec(SPEC)
sys.modules["to_environment_report"] = report_module
SPEC.loader.exec_module(report_module)


def make_settings(
    tmp_path: Path,
    inventory_path: Path,
    ndvr_folder: Path,
) -> dict:
    return {
        "files": {
            "default_input_file": str(inventory_path),
            "default_ndvr_file": str(ndvr_folder),
            "default_output_folder": str(tmp_path / "output"),
        },
        "required_columns": [
            "Release",
            "Project",
            "Element",
            "Type",
        ],
    }


def make_ndvr_line(
    element: str,
    type_: str,
    env: str,
    system: str,
    move_date: str,
    package: str,
    ccid: str = "CCID01",
) -> str:
    fields = [
        (element, 8),
        (type_, 8),
        (system, 8),
        ("SYS1", 4),
        (env, 5),
        (move_date, 10),
        ("12:00:00:00", 11),
        ("01.01", 5),
        ("USER01", 8),
        (ccid, 7),
        ("COMMENTS", 40),
        ("00004", 5),
        (package, 16),
    ]
    return " ".join(value.ljust(width)[:width] for value, width in fields)


def write_inventory(
    tmp_path: Path,
) -> Path:
    path = tmp_path / "inventory.xlsx"
    pd.DataFrame(
        [
            {
                "Release": "2026/07 release",
                "Project": "ABC12345",
                "Element": "PGM001",
                "Type": "OCOB",
            },
            {
                "Release": "2026/07 release",
                "Project": "CCID01LONG",
                "Element": "PGM002",
                "Type": "OCOB",
            },
        ]
    ).to_excel(path, index=False)
    return path


def write_ndvr_files(
    tmp_path: Path,
) -> Path:
    folder = tmp_path / "ndvr"
    folder.mkdir()
    old_file = folder / "NDVR-202607140800.txt"
    new_file = folder / "NDVR-202607141200.txt"
    old_file.write_text(
        "\n".join(
            [
                make_ndvr_line(
                    "PGM001",
                    "OCOB",
                    "QUAL1",
                    "PRIVATE1",
                    "2026/07/14",
                    "ABC",
                ),
                make_ndvr_line(
                    "UNTRACK",
                    "OCOB",
                    "PROD1",
                    "SHARED01",
                    "2026/07/14",
                    "ZZZ",
                ),
            ]
        ),
        encoding="cp1252",
    )
    new_file.write_text(
        make_ndvr_line(
            "PGM002",
            "OCOB",
            "PROD1",
            "PRIVATE1",
            "2026/07/14",
            "OTHER",
            ccid="CCID01",
        ),
        encoding="cp1252",
    )
    os.utime(old_file, (1_000, 1_000))
    os.utime(new_file, (2_000, 2_000))
    return folder


def test_to_environment_report_splits_moves_and_associates_inventory(
    tmp_path: Path,
) -> None:
    """TO QUAL/PROD reports split actual moves and explain inventory linkage."""
    inventory_path = write_inventory(tmp_path)
    ndvr_folder = write_ndvr_files(tmp_path)
    report = report_module.ToEnvironmentReport(
        settings=make_settings(tmp_path, inventory_path, ndvr_folder),
        base_dir=tmp_path,
    )

    rows_by_report = report.build_rows(date(2026, 7, 14))

    assert [row[5] for row in rows_by_report["QUAL"]] == ["PGM001"]
    assert rows_by_report["QUAL"][0][14] == "ABC12345"
    assert rows_by_report["QUAL"][0][15] == (
        "Linked by inventory project and NDVR package."
    )
    prod_associations = {row[5]: row[15] for row in rows_by_report["PROD"]}
    assert prod_associations == {
        "PGM002": "Assumption: linked based off CCID.",
        "UNTRACK": "Not associated to release.",
    }


def test_to_environment_report_writes_xlsx_and_pdf(
    tmp_path: Path,
) -> None:
    """Standalone TO reports write one XLSX and PDF for each target environment."""
    inventory_path = write_inventory(tmp_path)
    ndvr_folder = write_ndvr_files(tmp_path)
    report = report_module.ToEnvironmentReport(
        settings=make_settings(tmp_path, inventory_path, ndvr_folder),
        base_dir=tmp_path,
    )

    output_files = report.run(date(2026, 7, 14))

    assert {path.suffix for path in output_files} == {".xlsx", ".pdf"}
    assert len(output_files) == 4
    workbook = load_workbook(
        tmp_path / "output" / "TO QUAL" / "TO_QUAL_20260714.xlsx",
        read_only=True,
    )
    assert workbook.sheetnames == ["TO QUAL"]
    workbook.close()


def test_to_environment_report_can_write_xlsx_only(
    tmp_path: Path,
) -> None:
    """Verifies the all-report runner can suppress PDF output."""
    inventory_path = write_inventory(tmp_path)
    ndvr_folder = write_ndvr_files(tmp_path)
    report = report_module.ToEnvironmentReport(
        settings=make_settings(tmp_path, inventory_path, ndvr_folder),
        base_dir=tmp_path,
    )

    output_files = report.run(
        date(2026, 7, 14),
        formats=["xlsx"],
    )

    assert len(output_files) == 2
    assert {path.suffix for path in output_files} == {".xlsx"}
