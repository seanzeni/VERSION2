from __future__ import annotations

import importlib.util
import os
import sys
from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from app.core.models import ReleaseEffort


SCRIPT_PATH = Path(__file__).resolve().parents[1] / "scripts" / "ndvr_daily_move_audit.py"
SPEC = importlib.util.spec_from_file_location("ndvr_daily_move_audit", SCRIPT_PATH)
assert SPEC is not None and SPEC.loader is not None
audit_module = importlib.util.module_from_spec(SPEC)
sys.modules["ndvr_daily_move_audit"] = audit_module
SPEC.loader.exec_module(audit_module)


class FakeDbService:
    def get_efforts_for_release(
        self,
        release: str,
    ) -> list[ReleaseEffort]:
        return {
            "2026/07 release": [
                ReleaseEffort(
                    effort_id="ABC",
                    qual_date=date(2026, 7, 14),
                    prod_date=date(2026, 7, 15),
                )
            ],
            "2026/08 release": [
                ReleaseEffort(
                    effort_id="XYZ",
                    qual_date=date(2026, 8, 1),
                    prod_date=date(2026, 8, 2),
                )
            ],
        }.get(release, [])


def make_settings(
    tmp_path: Path,
    inventory_path: Path,
    ndvr_folder: Path,
) -> dict:
    return {
        "database": {},
        "files": {
            "default_input_file": str(inventory_path),
            "default_ndvr_file": str(ndvr_folder),
            "default_output_folder": str(tmp_path / "output"),
        },
        "required_columns": [
            "Release",
            "Project",
            "Element",
            "Type",
        ],
    }


def make_ndvr_line(
    element: str,
    type_: str,
    env: str,
    system: str,
    move_date: str,
    package: str,
) -> str:
    fields = [
        (element, 8),
        (type_, 8),
        (system, 8),
        ("SYS1", 4),
        (env, 5),
        (move_date, 10),
        ("12:00:00:00", 11),
        ("01.01", 5),
        ("USER01", 8),
        ("CCID01", 7),
        ("COMMENTS", 40),
        ("00004", 5),
        ("", 1),
        (package, 16),
    ]
    return " ".join(value.ljust(width)[:width] for value, width in fields)


def write_inventory(
    tmp_path: Path,
) -> Path:
    path = tmp_path / "inventory.xlsx"
    pd.DataFrame(
        [
            {
                "Release": "2026/07 release",
                "Project": "ABC",
                "Element": "PGM001",
                "Type": "OCOB",
            },
            {
                "Release": "2026/08 release",
                "Project": "XYZ",
                "Element": "PGM002",
                "Type": "OCOB",
            },
        ]
    ).to_excel(path, index=False)
    return path


def write_ndvr_files(
    tmp_path: Path,
) -> Path:
    folder = tmp_path / "ndvr"
    folder.mkdir()
    old_file = folder / "NDVR-202607140800.txt"
    new_file = folder / "NDVR-202607141200.txt"
    old_file.write_text(
        "\n".join(
            [
                make_ndvr_line(
                    "PGM001",
                    "OCOB",
                    "QUAL1",
                    "PRIVATE1",
                    "2026/07/14",
                    "PKG001",
                ),
                make_ndvr_line(
                    "NOINV01",
                    "OCOB",
                    "PROD1",
                    "SHARED01",
                    "2026/07/14",
                    "PKG002",
                ),
            ]
        ),
        encoding="cp1252",
    )
    new_file.write_text(
        "\n".join(
            [
                make_ndvr_line(
                    "PGM002",
                    "OCOB",
                    "QUAL1",
                    "PRIVATE1",
                    "2026/07/14",
                    "PKG003",
                ),
                make_ndvr_line(
                    "DROP001",
                    "OCOB",
                    "QUAL1",
                    "PUBLIC01",
                    "2026/07/14",
                    "PKG004",
                ),
            ]
        ),
        encoding="cp1252",
    )
    os.utime(old_file, (1_000, 1_000))
    os.utime(new_file, (2_000, 2_000))
    return folder


def test_daily_move_audit_builds_authorization_statuses(
    tmp_path: Path,
) -> None:
    """Verifies NDVR moves are classified against inventory and SQL dates."""
    inventory_path = write_inventory(tmp_path)
    ndvr_folder = write_ndvr_files(tmp_path)
    audit = audit_module.DailyMoveAudit(
        settings=make_settings(tmp_path, inventory_path, ndvr_folder),
        base_dir=tmp_path,
        db_service=FakeDbService(),
    )

    rows = audit.build_rows(date(2026, 7, 14))
    statuses = {row[6]: row[0] for row in rows}

    assert statuses == {
        "PGM001": "APPROVED_MOVE",
        "PGM002": "TRACKED_NOT_AUTHORIZED_FOR_DATE",
        "NOINV01": "NOT_TRACKED_IN_INVENTORY",
    }
    assert "2026-08-01" in next(row[13] for row in rows if row[6] == "PGM002")


def test_daily_move_audit_writes_xlsx_and_pdf(
    tmp_path: Path,
) -> None:
    """Verifies standalone audit output is XLSX and PDF only."""
    inventory_path = write_inventory(tmp_path)
    ndvr_folder = write_ndvr_files(tmp_path)
    audit = audit_module.DailyMoveAudit(
        settings=make_settings(tmp_path, inventory_path, ndvr_folder),
        base_dir=tmp_path,
        db_service=FakeDbService(),
    )

    output_files = audit.run(date(2026, 7, 14))

    assert {path.suffix for path in output_files} == {".xlsx", ".pdf"}
    workbook = load_workbook(output_files[0], read_only=True)
    assert workbook.sheetnames == ["Summary", "Detail"]
    workbook.close()
