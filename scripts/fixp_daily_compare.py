from __future__ import annotations

# Purpose:
#     Standalone day-over-day FIXP inventory comparison report.
#
# Usage:
#     py -3.14 scripts/fixp_daily_compare.py
#     py -3.14 scripts/fixp_daily_compare.py --date 2026-07-15

import argparse
import json
import re
import shlex
import subprocess
import sys
import tempfile
from collections import defaultdict
from dataclasses import dataclass
from datetime import date
from datetime import datetime
from datetime import timedelta
from pathlib import Path
from typing import Any
from urllib.parse import urlencode

import requests
import pyodbc
from openpyxl import load_workbook
from openpyxl.styles import Alignment

requests.packages.urllib3.disable_warnings(
    requests.packages.urllib3.exceptions.InsecureRequestWarning
)


REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from app.config.settings_loader import SettingsLoader  # noqa: E402
from app.core.models import MainframeLocationRecord  # noqa: E402
from app.core.release_rules import coerce_date  # noqa: E402
from app.reports.report_utils import export_xlsx  # noqa: E402
from app.reports.report_utils import make_read_only  # noqa: E402
from app.reports.report_utils import make_writable  # noqa: E402
from app.reports.report_utils import publish_staged_outputs  # noqa: E402
from app.reports.report_utils import report_date_stamp  # noqa: E402
from app.services.data_loader import DataLoader  # noqa: E402
from app.services.mainframe_location_service import MainframeLocationService  # noqa: E402


FIXP_FILE_PATTERN = re.compile(
    r"^FIXP-(?P<date>\d{8})_(?P<time>\d{6})\.txt$",
    re.IGNORECASE,
)
FIX_LIFECYCLE_ENVS = {"FIXT1", "FIXP1"}

DETAIL_HEADERS = [
    "Compare",
    "Stage",
    "System",
    "Subsys",
    "Element",
    "Type",
    "FIXP Date",
    "Days in FIXP",
    "FIXP CCID",
    "Inventory CCIDs",
    "Owner",
    "Manager",
    "Inventory",
    "Remarks",
    "DB_Issues_Fixes",
    "DB_Comments",
    "DB_Effort_ID",
    "DB_Owner",
    "DB_Manager",
    "DB_PROD_DATE",
]

OVERVIEW_HEADERS = [
    "Field",
    "Value",
]


@dataclass(frozen=True, slots=True)
class InventoryReference:
    release: str
    project: str
    team_lead: str
    team_lead_name: str = ""

    @property
    def label(self) -> str:
        return "-".join(
            value
            for value in (
                self.release,
                self.project,
                self.team_lead_name or self.team_lead,
            )
            if value
        )


@dataclass(frozen=True, slots=True)
class PersonDirectoryInfo:
    name: str
    employee_id: str = ""
    supervisor_id: str = ""


@dataclass(frozen=True, slots=True)
class OwnerManagerInfo:
    owner: str
    manager: str = ""


@dataclass(frozen=True, slots=True)
class FixpDatabaseReference:
    issues_fixes: str = ""
    comments: str = ""
    effort_id: str = ""
    owner: str = ""
    manager: str = ""
    prod_date: str = ""

    def as_columns(
        self,
    ) -> list[str]:
        return [
            self.issues_fixes,
            self.comments,
            self.effort_id,
            self.owner,
            self.manager,
            self.prod_date,
        ]


@dataclass(frozen=True, slots=True)
class FixpSnapshotRecord:
    record: MainframeLocationRecord
    file_timestamp: datetime


@dataclass(frozen=True, slots=True)
class FixpCompareDates:
    previous_date: date
    target_date: date


class FixpDailyCompare:
    def __init__(
        self,
        settings: dict[str, Any],
        base_dir: Path,
        fixp_source: Path | None = None,
        ndvr_source: Path | None = None,
        inventory_file: Path | None = None,
        output_folder: Path | None = None,
        fixp_database: Path | None = None,
        person_resolver=None,
        verbose: bool = False,
    ) -> None:
        self.settings = settings
        self.base_dir = base_dir
        fixp_source_value = fixp_source or settings["files"].get(
            "default_fixp_folder",
            "",
        )
        self.fixp_source = (
            self._resolve_path(fixp_source_value)
            if str(fixp_source_value).strip()
            else None
        )
        ndvr_source_value = ndvr_source or settings["files"].get(
            "default_ndvr_file",
            "",
        )
        self.ndvr_source = (
            self._resolve_path(ndvr_source_value)
            if str(ndvr_source_value).strip()
            else None
        )
        self.inventory_file = self._resolve_path(
            inventory_file or settings["files"]["default_input_file"]
        )
        self.output_folder = self._resolve_path(
            output_folder
            or settings["files"].get(
                "default_output_folder",
                "Output",
            )
        )
        fixp_database_value = fixp_database or settings["files"].get(
            "default_fixp_db",
            "",
        )
        self.fixp_database = (
            self._resolve_path(fixp_database_value)
            if str(fixp_database_value).strip()
            else None
        )
        self.person_resolver = person_resolver or PersonApiResolver(
            settings.get(
                "directory",
                {},
            ).get(
                "person_lookup_url",
                "",
            ),
            verbose=verbose,
        )

    def run(
        self,
        target_date: date | None,
    ) -> list[Path]:
        compare_dates = self._resolve_compare_dates(target_date)
        rows = self._build_rows(compare_dates)
        file_stem = f"FIXP_Daily_Stats_{report_date_stamp(compare_dates.target_date)}"
        self.output_folder.parent.mkdir(
            parents=True,
            exist_ok=True,
        )

        with tempfile.TemporaryDirectory(
            prefix="fixp-daily-",
            dir=self.output_folder.parent,
        ) as temp_dir:
            staged_path = Path(temp_dir) / f"{file_stem}.xlsx"
            export_xlsx(
                output_path=staged_path,
                sheets={
                    "Overview": (
                        OVERVIEW_HEADERS,
                        self._overview_rows(compare_dates),
                    ),
                    "FIXP Compare": (
                        DETAIL_HEADERS,
                        rows or self._empty_rows(compare_dates),
                    ),
                },
            )
            self._format_overview_sheet(staged_path)
            return publish_staged_outputs(
                staged_files=[staged_path],
                output_folder=self.output_folder,
                file_stems=[file_stem],
            )

    def build_rows(
        self,
        target_date: date | None,
    ) -> list[list[object]]:
        return self._build_rows(self._resolve_compare_dates(target_date))

    def _build_rows(
        self,
        compare_dates: FixpCompareDates,
    ) -> list[list[object]]:
        previous_snapshot = self._build_snapshot(compare_dates.previous_date)
        target_snapshot = self._build_snapshot(compare_dates.target_date)
        inventory_lookup = self._build_inventory_lookup()
        database_lookup = self._build_fixp_database_lookup()
        ndvr_service = self._load_latest_ndvr_service()

        rows: list[list[object]] = []
        target_lifecycle_keys = {
            self._lifecycle_record_key(snapshot_record.record)
            for snapshot_record in target_snapshot.values()
            if self._is_fix_lifecycle_record(snapshot_record.record)
        }
        previous_fixt_lifecycle_keys = {
            self._lifecycle_record_key(snapshot_record.record)
            for snapshot_record in previous_snapshot.values()
            if snapshot_record.record.env.strip().upper() == "FIXT1"
        }
        all_keys = sorted(
            set(previous_snapshot) | set(target_snapshot),
        )

        for key in all_keys:
            previous_record = previous_snapshot.get(key)
            target_record = target_snapshot.get(key)

            compare = self._compare(
                previous_record=previous_record,
                target_record=target_record,
            )
            if self._is_fixt_to_fixp_add(
                target_record=target_record,
                previous_fixt_lifecycle_keys=previous_fixt_lifecycle_keys,
            ):
                compare = "added"
            if compare == "deleted" and self._is_lifecycle_move_forward(
                previous_record=previous_record,
                target_lifecycle_keys=target_lifecycle_keys,
            ):
                continue

            display_record = (
                target_record.record
                if target_record is not None
                else previous_record.record
                if previous_record is not None
                else None
            )

            if display_record is None:
                continue

            inventory_references = inventory_lookup.get(display_record.key, [])
            owner_info = self._resolve_owner_info(
                user_id=display_record.user,
            )
            inventory_ccids = self._format_inventory_ccids(inventory_references)
            inventory = self._format_inventory(inventory_references)
            database_reference = database_lookup.get(
                self._database_record_key(display_record),
                FixpDatabaseReference(),
            )
            remarks = self._build_remarks(
                fixp_record=display_record,
                ndvr_service=ndvr_service,
            )
            rows.append(
                [
                    compare,
                    display_record.env,
                    display_record.system,
                    display_record.subsystem,
                    display_record.element,
                    display_record.type,
                    self._format_fixp_date(display_record.date_generated),
                    self._days_in_fixp(
                        fixp_record=display_record,
                        target_date=compare_dates.target_date,
                    ),
                    display_record.ccid,
                    inventory_ccids,
                    owner_info.owner,
                    owner_info.manager,
                    inventory,
                    remarks,
                    *database_reference.as_columns(),
                ]
            )

        return rows

    def _overview_rows(
        self,
        compare_dates: FixpCompareDates,
    ) -> list[list[object]]:
        previous_files = self._fixp_files_for_date(compare_dates.previous_date)
        target_files = self._fixp_files_for_date(compare_dates.target_date)

        return [
            [
                "Comparison",
                (
                    f"Comparing {compare_dates.target_date.isoformat()} to "
                    f"{compare_dates.previous_date.isoformat()} differences."
                ),
            ],
            [
                "Current Date",
                compare_dates.target_date.isoformat(),
            ],
            [
                "Compared Date",
                compare_dates.previous_date.isoformat(),
            ],
            [
                f"Files for {compare_dates.target_date.isoformat()}",
                self._format_file_list(target_files),
            ],
            [
                f"Files for {compare_dates.previous_date.isoformat()}",
                self._format_file_list(previous_files),
            ],
        ]

    def _resolve_compare_dates(
        self,
        target_date: date | None,
    ) -> FixpCompareDates:
        if target_date is not None:
            return FixpCompareDates(
                previous_date=target_date - timedelta(days=1),
                target_date=target_date,
            )

        available_dates = self._available_file_dates()
        if len(available_dates) < 2:
            raise FileNotFoundError(
                "At least two FIXP file dates are required when --date is not provided."
            )

        return FixpCompareDates(
            previous_date=available_dates[-2],
            target_date=available_dates[-1],
        )

    def _build_snapshot(
        self,
        target_date: date,
    ) -> dict[tuple[str, str, str, str, str], FixpSnapshotRecord]:
        snapshot: dict[tuple[str, str, str, str, str], FixpSnapshotRecord] = {}

        for file_path, file_timestamp in self._fixp_files_for_date(target_date):
            service = MainframeLocationService().load_file(file_path)
            for record in service.records:
                key = self._record_key(record)
                candidate = FixpSnapshotRecord(
                    record=record,
                    file_timestamp=file_timestamp,
                )
                existing = snapshot.get(key)

                if existing is None or self._is_older_daily_value(candidate, existing):
                    snapshot[key] = candidate

        return snapshot

    def _fixp_files_for_date(
        self,
        target_date: date,
    ) -> list[tuple[Path, datetime]]:
        files: list[tuple[Path, datetime]] = []
        for file_path in self._fixp_source_folder().glob("FIXP-*.txt"):
            file_timestamp = self._parse_file_timestamp(file_path)
            if file_timestamp is None or file_timestamp.date() != target_date:
                continue

            files.append(
                (
                    file_path,
                    file_timestamp,
                )
            )

        return sorted(
            files,
            key=lambda item: (
                item[1],
                item[0].name,
            ),
        )

    def _format_file_list(
        self,
        files: list[tuple[Path, datetime]],
    ) -> str:
        if not files:
            return "No files found."

        return "\n".join(file_path.name for file_path, _file_timestamp in files)

    def _format_overview_sheet(
        self,
        output_path: Path,
    ) -> None:
        make_writable(output_path)
        workbook = load_workbook(output_path)
        worksheet = workbook["Overview"]

        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True,
                )

        for row_index in (5, 6):
            value = str(worksheet.cell(row=row_index, column=2).value or "")
            line_count = max(
                value.count("\n") + 1,
                1,
            )
            worksheet.row_dimensions[row_index].height = max(
                18,
                line_count * 15,
            )

        worksheet.column_dimensions["A"].width = 28
        worksheet.column_dimensions["B"].width = 60
        workbook.save(output_path)
        make_read_only(output_path)

    def _available_file_dates(
        self,
    ) -> list[date]:
        return sorted(
            {
                file_timestamp.date()
                for file_path in self._fixp_source_folder().glob("FIXP-*.txt")
                if (file_timestamp := self._parse_file_timestamp(file_path)) is not None
            }
        )

    def _fixp_source_folder(
        self,
    ) -> Path:
        source = self.fixp_source
        if source is None:
            raise FileNotFoundError(
                "FIXP source folder was not configured. Set files.default_fixp_folder "
                "or pass --fixp-source."
            )

        folder = source.parent if source.is_file() else source

        if not folder.exists():
            raise FileNotFoundError(f"FIXP source was not found: {folder}")

        if not folder.is_dir():
            raise NotADirectoryError(f"FIXP source is not a directory: {folder}")

        return folder

    def _load_latest_ndvr_service(
        self,
    ) -> MainframeLocationService | None:
        latest_file = self._latest_ndvr_file()
        if latest_file is None:
            return None

        return MainframeLocationService().load_file(latest_file)

    def _latest_ndvr_file(
        self,
    ) -> Path | None:
        source = self.ndvr_source
        if source is None:
            return None

        if source.is_file():
            return source

        if not source.exists() or not source.is_dir():
            return None

        files = [
            file_path
            for pattern in ("*.txt", "*.dat", "*.csv")
            for file_path in source.glob(pattern)
            if file_path.is_file()
        ]

        if not files:
            return None

        return max(
            files,
            key=lambda file_path: (
                file_path.stat().st_mtime,
                file_path.name,
            ),
        )

    def _parse_file_timestamp(
        self,
        file_path: Path,
    ) -> datetime | None:
        match = FIXP_FILE_PATTERN.match(file_path.name)
        if match is None:
            return None

        return datetime.strptime(
            f"{match.group('date')}{match.group('time')}",
            "%Y%m%d%H%M%S",
        )

    def _build_inventory_lookup(
        self,
    ) -> dict[tuple[str, str], list[InventoryReference]]:
        data_loader = DataLoader(
            file_path=self.inventory_file,
            required_columns=self.settings["required_columns"],
        )
        dataframe = data_loader.load()
        lookup: dict[tuple[str, str], list[InventoryReference]] = defaultdict(list)

        for _, row in dataframe.iterrows():
            element = str(row.get("Element", "")).strip().upper()
            type_ = str(row.get("Type", "")).strip().upper()

            if not element or not type_:
                continue

            lookup[
                (
                    element,
                    type_,
                )
            ].append(self._build_inventory_reference(row))

        return dict(lookup)

    def _build_inventory_reference(
        self,
        row,
    ) -> InventoryReference:
        team_lead = str(row.get("DSN ID", "")).strip()[:4]
        team_lead_name = (
            self.person_resolver.resolve(team_lead).name if team_lead else ""
        )

        return InventoryReference(
            release=str(row.get("Release", "")).strip(),
            project=str(row.get("Project", "")).strip(),
            team_lead=team_lead,
            team_lead_name=team_lead_name,
        )

    def _build_fixp_database_lookup(
        self,
    ) -> dict[tuple[str, str, str, str], FixpDatabaseReference]:
        if self.fixp_database is None:
            return {}

        return AccessFixpReferenceLoader(
            database_path=self.fixp_database,
            fallback_python=self.settings.get(
                "files",
                {},
            ).get(
                "fixp_32bit_python",
                "py -3.14-32",
            ),
        ).load()

    def _compare(
        self,
        previous_record: FixpSnapshotRecord | None,
        target_record: FixpSnapshotRecord | None,
    ) -> str:
        if target_record is None:
            return "deleted"

        if previous_record is None:
            return "added"

        if self._fixp_generated_date(
            previous_record.record
        ) == self._fixp_generated_date(target_record.record):
            return "no change"

        return "modified"

    def _record_key(
        self,
        record: MainframeLocationRecord,
    ) -> tuple[str, str, str, str, str]:
        return (
            record.env.strip().upper(),
            record.system.strip().upper(),
            record.subsystem.strip().upper(),
            record.element.strip().upper(),
            record.type.strip().upper(),
        )

    def _lifecycle_record_key(
        self,
        record: MainframeLocationRecord,
    ) -> tuple[str, str, str, str]:
        return (
            record.system.strip().upper(),
            record.subsystem.strip().upper(),
            record.element.strip().upper(),
            record.type.strip().upper(),
        )

    def _is_fix_lifecycle_record(
        self,
        record: MainframeLocationRecord,
    ) -> bool:
        return record.env.strip().upper() in FIX_LIFECYCLE_ENVS

    def _is_lifecycle_move_forward(
        self,
        previous_record: FixpSnapshotRecord | None,
        target_lifecycle_keys: set[tuple[str, str, str, str]],
    ) -> bool:
        if previous_record is None:
            return False

        record = previous_record.record
        return (
            self._is_fix_lifecycle_record(record)
            and self._lifecycle_record_key(record) in target_lifecycle_keys
        )

    def _is_fixt_to_fixp_add(
        self,
        target_record: FixpSnapshotRecord | None,
        previous_fixt_lifecycle_keys: set[tuple[str, str, str, str]],
    ) -> bool:
        if target_record is None:
            return False

        record = target_record.record
        return (
            record.env.strip().upper() == "FIXP1"
            and self._lifecycle_record_key(record) in previous_fixt_lifecycle_keys
        )

    def _database_record_key(
        self,
        record: MainframeLocationRecord,
    ) -> tuple[str, str, str, str]:
        return (
            record.element.strip().upper(),
            record.type.strip().upper(),
            record.system.strip().upper(),
            record.subsystem.strip().upper(),
        )

    def _is_older_daily_value(
        self,
        candidate: FixpSnapshotRecord,
        existing: FixpSnapshotRecord,
    ) -> bool:
        candidate_date = self._fixp_generated_date(candidate.record) or date.min
        existing_date = self._fixp_generated_date(existing.record) or date.min

        return (
            candidate_date,
            candidate.file_timestamp,
            candidate.record.time_generated,
        ) < (
            existing_date,
            existing.file_timestamp,
            existing.record.time_generated,
        )

    def _format_fixp_date(
        self,
        value: str,
    ) -> str:
        parsed_date = coerce_date(value)
        if parsed_date is None:
            return str(value).strip()

        return parsed_date.strftime("%d-%b-%y")

    def _fixp_generated_date(
        self,
        record: MainframeLocationRecord,
    ) -> date | None:
        return coerce_date(record.date_generated)

    def _days_in_fixp(
        self,
        fixp_record: MainframeLocationRecord,
        target_date: date,
    ) -> int | str:
        source_date = coerce_date(fixp_record.source_date)
        generated_date = self._fixp_generated_date(fixp_record)
        if source_date is not None and generated_date is not None:
            return max(
                (generated_date - source_date).days,
                0,
            )

        if generated_date is not None:
            return max(
                (target_date - generated_date).days,
                0,
            )

        if source_date is None:
            return ""

        return max(
            (target_date - source_date).days,
            0,
        )

    def _format_inventory(
        self,
        references: list[InventoryReference],
    ) -> str:
        return "; ".join(
            sorted({reference.label for reference in references if reference.label})
        )

    def _format_inventory_ccids(
        self,
        references: list[InventoryReference],
    ) -> str:
        return "; ".join(
            sorted({reference.project for reference in references if reference.project})
        )

    def _build_remarks(
        self,
        fixp_record: MainframeLocationRecord,
        ndvr_service: MainframeLocationService | None,
    ) -> str:
        if ndvr_service is None:
            return ""

        fixp_date = self._fixp_generated_date(fixp_record)
        if fixp_date is None:
            return ""

        has_newer_prod = any(
            record.env.strip().upper() == "PROD1"
            and (prod_date := coerce_date(record.date_generated)) is not None
            and prod_date > fixp_date
            for record in ndvr_service.find(
                element=fixp_record.element,
                type_=fixp_record.type,
            )
        )

        if has_newer_prod:
            return "Newer version in PROD"

        return ""

    def _resolve_owner_info(
        self,
        user_id: str,
    ) -> OwnerManagerInfo:
        owner = self.person_resolver.resolve(user_id)
        manager = (
            self.person_resolver.resolve_name(owner.supervisor_id)
            if owner.supervisor_id
            else ""
        )

        return OwnerManagerInfo(
            owner=owner.name,
            manager=manager,
        )

    def _empty_rows(
        self,
        compare_dates: FixpCompareDates,
    ) -> list[list[object]]:
        return [
            [
                "no change",
                "",
                "",
                "",
                "",
                "",
                compare_dates.target_date.strftime("%d-%b-%y"),
                "",
                "",
                "",
                "",
                "",
                (
                    "No FIXP records found between "
                    f"{compare_dates.previous_date.isoformat()} and "
                    f"{compare_dates.target_date.isoformat()}."
                ),
                "",
                "",
                "",
                "",
                "",
                "",
                "",
            ]
        ]

    def _resolve_path(
        self,
        value: str | Path,
    ) -> Path:
        path = Path(value)
        if path.is_absolute():
            return path
        return self.base_dir / path


class AccessFixpReferenceLoader:
    TABLE_NAME = "tblFIXP1"

    def __init__(
        self,
        database_path: Path,
        fallback_python: str = "py -3.14-32",
        allow_fallback: bool = True,
    ) -> None:
        self.database_path = Path(database_path)
        self.fallback_python = str(fallback_python).strip()
        self.allow_fallback = allow_fallback

    def load(
        self,
    ) -> dict[tuple[str, str, str, str], FixpDatabaseReference]:
        try:
            return self._load_with_pyodbc()
        except pyodbc.Error as exc:
            if not self.allow_fallback or not self._is_missing_access_driver(exc):
                raise

            return self._load_with_32bit_python()

    def _load_with_pyodbc(
        self,
    ) -> dict[tuple[str, str, str, str], FixpDatabaseReference]:
        if not self.database_path.exists():
            raise FileNotFoundError(
                f"FIXP Access database was not found: {self.database_path}"
            )

        query = f"SELECT * FROM [{self.TABLE_NAME}]"
        lookup: dict[tuple[str, str, str, str], FixpDatabaseReference] = {}

        with pyodbc.connect(self._connection_string()) as connection:
            cursor = connection.cursor()
            cursor.execute(query)
            column_names = [str(column[0]) for column in cursor.description]

            for row in cursor.fetchall():
                row_data = {
                    self._normalize_column_name(column_name): value
                    for column_name, value in zip(column_names, row, strict=False)
                }
                key = self._row_key(row_data)
                if key is None:
                    continue

                lookup[key] = FixpDatabaseReference(
                    issues_fixes=self._clean(row_data.get("issuesfixes", "")),
                    comments=self._clean(row_data.get("comments", "")),
                    effort_id=self._clean(row_data.get("effortid", "")),
                    owner=self._clean(row_data.get("owner", "")),
                    manager=self._clean(row_data.get("manager", "")),
                    prod_date=self._clean(row_data.get("proddate", "")),
                )

        return lookup

    def _load_with_32bit_python(
        self,
    ) -> dict[tuple[str, str, str, str], FixpDatabaseReference]:
        if not self.fallback_python:
            raise RuntimeError(
                "The Microsoft Access ODBC driver is not available to this Python, "
                "and files.fixp_32bit_python is not configured."
            )

        command = [
            *shlex.split(
                self.fallback_python,
                posix=False,
            ),
            str(Path(__file__).with_name("fixp_access_dump.py").resolve()),
            str(self.database_path),
        ]
        result = subprocess.run(
            command,
            capture_output=True,
            check=False,
            encoding="utf-8",
            errors="replace",
            timeout=60,
        )

        if result.returncode != 0:
            error = result.stderr.strip() or result.stdout.strip() or "no output"
            raise RuntimeError(
                "Unable to read FIXP Access database using the configured 32-bit "
                f"Python command [{self.fallback_python}]: {error}"
            )

        try:
            payload = json.loads(result.stdout)
        except json.JSONDecodeError as exc:
            raise RuntimeError(
                "The 32-bit FIXP Access database helper returned invalid JSON."
            ) from exc

        return self._lookup_from_payload(payload)

    def dump_json(
        self,
    ) -> str:
        payload = [
            {
                "element": key[0],
                "type": key[1],
                "system": key[2],
                "subsystem": key[3],
                "issues_fixes": reference.issues_fixes,
                "comments": reference.comments,
                "effort_id": reference.effort_id,
                "owner": reference.owner,
                "manager": reference.manager,
                "prod_date": reference.prod_date,
            }
            for key, reference in self._load_with_pyodbc().items()
        ]
        return json.dumps(payload)

    def _lookup_from_payload(
        self,
        payload,
    ) -> dict[tuple[str, str, str, str], FixpDatabaseReference]:
        lookup: dict[tuple[str, str, str, str], FixpDatabaseReference] = {}
        if not isinstance(payload, list):
            raise RuntimeError(
                "The 32-bit FIXP Access database helper returned non-list JSON."
            )

        for item in payload:
            if not isinstance(item, dict):
                continue

            key = self._row_key(item)
            if key is None:
                continue

            lookup[key] = FixpDatabaseReference(
                issues_fixes=self._clean(item.get("issues_fixes", "")),
                comments=self._clean(item.get("comments", "")),
                effort_id=self._clean(item.get("effort_id", "")),
                owner=self._clean(item.get("owner", "")),
                manager=self._clean(item.get("manager", "")),
                prod_date=self._clean(item.get("prod_date", "")),
            )

        return lookup

    def _is_missing_access_driver(
        self,
        exc: pyodbc.Error,
    ) -> bool:
        message = str(exc).upper()
        return "IM001" in message or "DATA SOURCE NAME NOT FOUND" in message

    def _connection_string(
        self,
    ) -> str:
        return (
            "DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};"
            f"DBQ={self.database_path};"
        )

    def _row_key(
        self,
        row_data: dict[str, object],
    ) -> tuple[str, str, str, str] | None:
        element = self._clean(row_data.get("element", "")).upper()
        type_ = self._clean(row_data.get("type", "")).upper()
        system = self._clean(row_data.get("system", "")).upper()
        subsystem = self._clean(
            row_data.get(
                "subsystem",
                row_data.get(
                    "subsytem",
                    "",
                ),
            )
        ).upper()

        if not all(
            (
                element,
                type_,
                system,
                subsystem,
            )
        ):
            return None

        return (
            element,
            type_,
            system,
            subsystem,
        )

    def _normalize_column_name(
        self,
        value: str,
    ) -> str:
        return re.sub(
            r"[^a-z0-9]",
            "",
            str(value).strip().lower(),
        )

    def _clean(
        self,
        value: object,
    ) -> str:
        if value is None:
            return ""

        if isinstance(value, (date, datetime)):
            return value.strftime("%Y-%m-%d")

        return str(value).strip()


class PersonApiResolver:
    def __init__(
        self,
        lookup_url: str,
        name_resolver=None,
        verbose: bool = False,
    ) -> None:
        self.lookup_url = str(lookup_url).strip()
        self.verbose = verbose
        self.name_resolver = name_resolver or ActiveDirectoryNameResolver(
            verbose=verbose,
        )
        self._cache: dict[str, PersonDirectoryInfo] = {}
        self._debug(
            f"Person lookup URL configured: {'yes' if self.lookup_url else 'no'}"
        )

    def resolve(
        self,
        criteria: str,
    ) -> PersonDirectoryInfo:
        clean_criteria = str(criteria).strip()
        if not clean_criteria:
            return PersonDirectoryInfo(name="")

        cache_key = clean_criteria.upper()
        if cache_key not in self._cache:
            self._cache[cache_key] = self._lookup(clean_criteria)

        return self._cache[cache_key]

    def _lookup(
        self,
        criteria: str,
    ) -> PersonDirectoryInfo:
        if not self.lookup_url:
            self._debug(
                f"API lookup skipped for {criteria!r}; no person_lookup_url configured."
            )
            return PersonDirectoryInfo(
                name=self.resolve_name(criteria),
            )

        request_url = self._lookup_request_url(criteria)
        self._debug(f"API lookup for {criteria!r}: {request_url}")

        payload = self._request_payload(
            criteria=criteria,
            request_url=request_url,
        )
        if payload is None:
            return PersonDirectoryInfo(name=criteria)

        self._debug(f"API payload for {criteria!r}: {self._payload_summary(payload)}")
        person = self._extract_person_payload(payload)
        if not person:
            self._debug(f"API lookup found no person object for {criteria!r}.")
            return PersonDirectoryInfo(name=criteria)

        employee_id = self._first_value(
            person,
            (
                "employeeId",
                "employeeID",
                "adId",
                "adID",
                "id",
                "networkId",
                "networkID",
                "samAccountName",
                "sAMAccountName",
                "userId",
                "userID",
            ),
        )
        supervisor_id = self._first_value(
            person,
            (
                "supervisorId",
                "supervisorID",
                "managerId",
                "managerID",
                "supervisorAdId",
                "supervisorADID",
                "supervisorUserId",
                "supervisorUserID",
                "supervisorEmployeeId",
            ),
        )
        self._debug(
            f"API extracted for {criteria!r}: employee_id={employee_id!r}, "
            f"supervisor_id={supervisor_id!r}"
        )

        return PersonDirectoryInfo(
            name=self.resolve_name(employee_id or criteria),
            employee_id=employee_id,
            supervisor_id=supervisor_id,
        )

    def _request_payload(
        self,
        criteria: str,
        request_url: str,
    ):
        try:
            response = requests.get(
                request_url,
                timeout=10,
                verify=False,
            )
            response.raise_for_status()
            return response.json()
        except requests.exceptions.HTTPError as exc:
            self._debug(
                f"Requests API lookup failed for {criteria!r}: HTTP "
                f"{exc.response.status_code if exc.response is not None else 'unknown'}"
            )
        except requests.exceptions.SSLError as exc:
            self._debug(
                f"Requests SSL issue ignored for {criteria!r}; "
                f"falling back to PowerShell: {exc}"
            )
        except requests.exceptions.RequestException as exc:
            self._debug(
                f"Requests API lookup failed for {criteria!r}: "
                f"{type(exc).__name__}: {exc}"
            )
        except ValueError as exc:
            self._debug(
                f"Requests API lookup failed for {criteria!r}: invalid JSON {exc}"
            )

        return self._request_payload_with_powershell(
            criteria=criteria,
            request_url=request_url,
        )

    def _request_payload_with_powershell(
        self,
        criteria: str,
        request_url: str,
    ):
        self._debug(f"Trying PowerShell API lookup for {criteria!r}.")

        try:
            result = subprocess.run(
                [
                    "powershell.exe",
                    "-NoProfile",
                    "-NonInteractive",
                    "-Command",
                    self._build_api_lookup_script(request_url),
                ],
                capture_output=True,
                check=False,
                encoding="utf-8",
                errors="replace",
                timeout=15,
            )
        except (OSError, subprocess.TimeoutExpired) as exc:
            self._debug(
                f"PowerShell API lookup failed for {criteria!r}: "
                f"{type(exc).__name__}: {exc}"
            )
            return None

        if result.returncode != 0 or not result.stdout.strip():
            error = result.stderr.strip() if result.stderr else "no output"
            self._debug(f"PowerShell API lookup failed for {criteria!r}: {error}")
            return None

        try:
            return json.loads(result.stdout)
        except json.JSONDecodeError as exc:
            self._debug(
                f"PowerShell API lookup failed for {criteria!r}: invalid JSON {exc}"
            )
            return None

    def _build_api_lookup_script(
        self,
        request_url: str,
    ) -> str:
        escaped_url = request_url.replace(
            "'",
            "''",
        )
        return (
            "$ErrorActionPreference = 'Stop'; "
            f"$response = Invoke-RestMethod -Method Get -Uri '{escaped_url}' "
            "-UseDefaultCredentials; "
            "$response | ConvertTo-Json -Depth 10 -Compress"
        )

    def resolve_name(
        self,
        ad_id: str,
    ) -> str:
        clean_ad_id = str(ad_id).strip()
        if not clean_ad_id:
            return ""

        return self.name_resolver.resolve_name(clean_ad_id)

    def _debug(
        self,
        message: str,
    ) -> None:
        if self.verbose:
            print(
                f"[fixp-directory] {message}",
                file=sys.stderr,
            )

    def _lookup_request_url(
        self,
        criteria: str,
    ) -> str:
        if "{criteria}" in self.lookup_url:
            return self.lookup_url.replace(
                "{criteria}",
                urlencode({"": criteria})[1:],
            )

        for placeholder in (
            "xxxxx",
            "XXXXX",
        ):
            if placeholder in self.lookup_url:
                return self.lookup_url.replace(
                    placeholder,
                    urlencode({"": criteria})[1:],
                )

        if self.lookup_url.endswith(("criteria=", "criteria%3D")):
            return f"{self.lookup_url}{urlencode({'': criteria})[1:]}"

        separator = "&" if "?" in self.lookup_url else "?"
        return f"{self.lookup_url}{separator}{urlencode({'criteria': criteria})}"

    def _extract_person_payload(
        self,
        payload,
    ) -> dict[str, Any]:
        if isinstance(payload, list):
            return self._extract_person_payload(payload[0]) if payload else {}

        if not isinstance(payload, dict):
            return {}

        for key in (
            "data",
            "result",
            "results",
            "items",
            "value",
        ):
            value = payload.get(key)
            if isinstance(value, (dict, list)):
                extracted = self._extract_person_payload(value)
                if extracted:
                    return extracted

        return payload

    def _payload_summary(
        self,
        payload,
    ) -> str:
        if isinstance(payload, list):
            return f"list(len={len(payload)})"

        if isinstance(payload, dict):
            keys = ", ".join(sorted(str(key) for key in payload)[:20])
            return f"dict(keys=[{keys}])"

        return type(payload).__name__

    def _first_value(
        self,
        payload: dict[str, Any],
        keys: tuple[str, ...],
    ) -> str:
        normalized = {str(key).lower(): value for key, value in payload.items()}

        for key in keys:
            value = normalized.get(key.lower())
            if value is not None and str(value).strip():
                return str(value).strip()

        return ""


class ActiveDirectoryNameResolver:
    def __init__(
        self,
        verbose: bool = False,
    ) -> None:
        self.verbose = verbose
        self._cache: dict[str, str] = {}

    def resolve_name(
        self,
        ad_id: str,
    ) -> str:
        clean_ad_id = str(ad_id).strip()
        if not clean_ad_id:
            return ""

        cache_key = clean_ad_id.upper()
        if cache_key not in self._cache:
            self._cache[cache_key] = self._lookup_name(clean_ad_id)

        return self._cache[cache_key]

    def _lookup_name(
        self,
        ad_id: str,
    ) -> str:
        try:
            self._debug(f"AD lookup for {ad_id!r}.")
            result = subprocess.run(
                [
                    "powershell.exe",
                    "-NoProfile",
                    "-NonInteractive",
                    "-Command",
                    self._build_lookup_script(ad_id),
                ],
                capture_output=True,
                check=False,
                encoding="utf-8",
                errors="replace",
                timeout=10,
            )
        except (OSError, subprocess.TimeoutExpired):
            self._debug(f"AD lookup failed for {ad_id!r}: process error or timeout.")
            return ad_id

        if result.returncode != 0 or not result.stdout.strip():
            error = result.stderr.strip() if result.stderr else "no output"
            self._debug(f"AD lookup failed for {ad_id!r}: {error}")
            return ad_id

        try:
            payload = json.loads(result.stdout)
        except json.JSONDecodeError:
            self._debug(f"AD lookup failed for {ad_id!r}: invalid JSON.")
            return ad_id

        display_name = str(payload.get("DisplayName") or ad_id).strip()
        self._debug(f"AD lookup result for {ad_id!r}: {display_name!r}")
        return display_name

    def _debug(
        self,
        message: str,
    ) -> None:
        if self.verbose:
            print(
                f"[fixp-directory] {message}",
                file=sys.stderr,
            )

    def _build_lookup_script(
        self,
        ad_id: str,
    ) -> str:
        escaped_ad_id = ad_id.replace(
            "'",
            "''",
        )
        return (
            "$ErrorActionPreference = 'Stop'; "
            "Import-Module ActiveDirectory; "
            f"$lookup = '{escaped_ad_id}'; "
            "$user = $null; "
            "try { "
            "$user = Get-ADUser -Identity $lookup -Properties DisplayName "
            "} catch { "
            '$safeLookup = $lookup.Replace("\'", "\'\'"); '
            "$user = Get-ADUser "
            "-Filter \"EmployeeID -eq '$safeLookup' -or SamAccountName -eq '$safeLookup'\" "
            "-Properties DisplayName | Select-Object -First 1 "
            "}; "
            "[PSCustomObject]@{ "
            "DisplayName = $user.DisplayName "
            "} | ConvertTo-Json -Compress"
        )


def parse_args(
    argv: list[str] | None = None,
) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Create a day-over-day FIXP comparison report."
    )
    parser.add_argument(
        "--settings",
        default=str(REPO_ROOT / "settings.json"),
        help="Path to settings.json. Defaults to the repository settings file.",
    )
    parser.add_argument(
        "--date",
        help=(
            "Report date in YYYY-MM-DD format. Defaults to the latest two FIXP "
            "file dates available."
        ),
    )
    parser.add_argument(
        "--fixp-source",
        help="Optional FIXP source directory. Defaults to files.default_fixp_folder.",
    )
    parser.add_argument(
        "--ndvr-source",
        help="Optional NDVR inventory source directory or file. Defaults to settings.",
    )
    parser.add_argument(
        "--inventory-file",
        help="Optional inventory spreadsheet path. Defaults to settings.",
    )
    parser.add_argument(
        "--output-folder",
        help="Optional output folder. Defaults to settings.",
    )
    parser.add_argument(
        "--fixp-db",
        help=(
            "Optional Access database path containing tblFIXP1. "
            "Defaults to files.default_fixp_db when configured."
        ),
    )
    parser.add_argument(
        "--dump-fixp-db-json",
        help=argparse.SUPPRESS,
    )
    parser.add_argument(
        "--verbose",
        action="store_true",
        help="Print person API and AD lookup diagnostics to stderr.",
    )
    return parser.parse_args(argv)


def parse_target_date(
    value: str | None,
    today: date | None = None,
) -> date | None:
    if not value:
        return None

    return datetime.strptime(value, "%Y-%m-%d").date()


def main(
    argv: list[str] | None = None,
) -> int:
    args = parse_args(argv)
    if args.dump_fixp_db_json:
        print(
            AccessFixpReferenceLoader(
                database_path=Path(args.dump_fixp_db_json),
                allow_fallback=False,
            ).dump_json()
        )
        return 0

    settings_path = Path(args.settings).resolve()
    settings = SettingsLoader(settings_path).load()
    base_dir = settings_path.parent
    target_date = parse_target_date(args.date)

    generated_files = FixpDailyCompare(
        settings=settings,
        base_dir=base_dir,
        fixp_source=Path(args.fixp_source) if args.fixp_source else None,
        ndvr_source=Path(args.ndvr_source) if args.ndvr_source else None,
        inventory_file=Path(args.inventory_file) if args.inventory_file else None,
        output_folder=Path(args.output_folder) if args.output_folder else None,
        fixp_database=Path(args.fixp_db) if args.fixp_db else None,
        verbose=args.verbose,
    ).run(target_date)

    print("Generated:")
    for file_path in generated_files:
        print(f"- {file_path}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
